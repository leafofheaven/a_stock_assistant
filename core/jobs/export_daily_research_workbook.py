"""Export a read-only daily research workbook as an Excel file."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import Settings, get_settings
from core.advice.simulated_trading_advice import build_simulated_trading_advice, summarize_simulated_trading_advice
from core.diagnostics.data_quality_snapshot import build_data_quality_snapshot
from core.storage.duckdb_store import DuckDBStore, DuckDBStoreError
from core.technical.elder import build_elder_review

SHEET_NAMES = [
    "00_摘要",
    "01_今日候选",
    "03_买入区间",
    "04_观察池",
    "06_外部模拟持仓",
    "13_模拟交易建议",
]
SELECTION_DISPLAY_TOP_N = 10
WATCHLIST_DISPLAY_TOP_N = 30

SENSITIVE_KEYWORDS = ("token", "key", "password", "secret", "api密钥", "smtp")
SENSITIVE_VALUE_MARKERS = ("sk-",)
PROJECT_ROOT = Path(__file__).resolve().parents[2]
FUNDAMENTAL_SCORE_MISSING_NOTE_PREFIX = "基本面分（fundamental_score）缺失记录数"
ELDER_DISPLAY_COLUMNS = [
    "elder_score",
    "action_hint",
    "elder_reason",
    "weekly_trend",
    "daily_pullback",
    "force_signal",
    "elder_ray_signal",
]

RANK_COLUMNS = {
    "rank",
    "candidate_rank",
    "original_rank",
    "today_rank",
    "previous_rank",
    "rank_change",
    "best_rank",
    "latest_rank",
}

COLUMN_LABELS = {
    "display_order": "序号",
    "message": "说明",
    "metric": "项目",
    "value": "内容",
    "section": "项目",
    "description": "说明",
    "source": "来源（source）",
    "source_tags": "来源标签（source_tags）",
    "trade_date": "交易日期（trade_date）",
    "review_date": "复核日期（review_date）",
    "latest_trade_date": "最新行情日期（latest_trade_date）",
    "snapshot_date": "快照日期（snapshot_date）",
    "ts_code": "股票代码（ts_code）",
    "symbol": "股票简称代码（symbol）",
    "name": "股票名称（name）",
    "industry": "行业（industry）",
    "market": "市场（market）",
    "list_date": "上市日期（list_date）",
    "close": "收盘价（close）",
    "current_close": "当前价格（current_close）",
    "current_price": "当前价格（current_price）",
    "pe": "市盈率（pe）",
    "pb": "市净率（pb）",
    "total_score": "综合分（total_score）",
    "latest_score": "最近综合分（latest_score）",
    "trend_score": "趋势分（trend_score）",
    "momentum_score": "动量分（momentum_score）",
    "liquidity_score": "流动性分（liquidity_score）",
    "fundamental_score": "基本面分（fundamental_score）",
    "volatility_score": "波动分（volatility_score）",
    "quality_score": "质量分（quality_score）",
    "valuation_score": "估值分（valuation_score）",
    "risk_score": "风险分（risk_score）",
    "select_reason": "候选原因（select_reason）",
    "risk_note": "风险提示（risk_note）",
    "holding_status": "持仓状态（holding_status）",
    "simulated_action": "模拟操作建议（simulated_action）",
    "suggested_position": "建议模拟仓位（suggested_position）",
    "action_priority": "动作优先级（action_priority）",
    "elder_score": "埃尔德分（elder_score）",
    "review_scope": "复核范围（review_scope）",
    "review_status": "复核状态（review_status）",
    "review_reason": "复核说明（review_reason）",
    "action_hint": "操作提示（action_hint）",
    "review_action": "复核动作（review_action）",
    "elder_reason": "复核原因（elder_reason）",
    "weekly_trend": "周线趋势（weekly_trend）",
    "daily_pullback": "日线回调（daily_pullback）",
    "force_signal": "强力指数信号（force_signal）",
    "elder_ray_signal": "埃尔德射线信号（elder_ray_signal）",
    "ema13": "13日指数移动平均线（ema13）",
    "ema22": "22日指数移动平均线（ema22）",
    "ema60": "60日指数移动平均线（ema60）",
    "support_20d": "20日支撑位（support_20d）",
    "support_60d": "60日支撑位（support_60d）",
    "resistance_20d": "20日阻力位（resistance_20d）",
    "resistance_60d": "60日阻力位（resistance_60d）",
    "nearest_support": "最近支撑位（nearest_support）",
    "nearest_resistance": "最近阻力位（nearest_resistance）",
    "atr_14": "14日平均真实波幅（atr_14）",
    "entry_low": "买入区间下限（entry_low）",
    "entry_high": "买入区间上限（entry_high）",
    "entry_mid": "买入区间中值（entry_mid）",
    "stop_loss": "止损价（stop_loss）",
    "target_price": "目标价（target_price）",
    "risk_pct": "风险比例（risk_pct）",
    "reward_pct": "收益比例（reward_pct）",
    "reward_risk_ratio": "盈亏比（reward_risk_ratio）",
    "entry_zone_status": "买入区间状态（entry_zone_status）",
    "entry_zone_status_cn": "买入区间状态说明（entry_zone_status_cn）",
    "missing_reason": "缺失原因（missing_reason）",
    "chase_risk": "追高风险（chase_risk）",
    "chase_risk_cn": "追高风险说明（chase_risk_cn）",
    "price_action_note": "价格行为说明（price_action_note）",
    "entry_reason": "入池原因（entry_reason）",
    "watch_reason": "观察原因（watch_reason）",
    "watch_status": "观察状态（watch_status）",
    "watch_status_label": "观察状态说明（watch_status_label）",
    "watch_days": "观察天数（watch_days）",
    "first_selected_date": "首次入选日期（first_selected_date）",
    "last_selected_date": "最近入选日期（last_selected_date）",
    "selected_count_5d": "近5日入选次数（selected_count_5d）",
    "selected_count_10d": "近10日入选次数（selected_count_10d）",
    "consecutive_selected_days": "连续入选天数（consecutive_selected_days）",
    "total_score_change": "综合分变化（total_score_change）",
    "score_change": "综合分变化（score_change）",
    "is_top_n": "是否进入前N名（is_top_n）",
    "top_n_flag": "是否进入前N名（top_n_flag）",
    "is_new_candidate": "是否新候选（is_new_candidate）",
    "new_candidate_flag": "是否新候选（new_candidate_flag）",
    "daily_note": "每日备注（daily_note）",
    "platform": "平台（platform）",
    "account_name": "账户名称（account_name）",
    "quantity": "持仓数量（quantity）",
    "cost_price": "持仓成本（cost_price）",
    "market_value": "市值（market_value）",
    "pnl": "浮动盈亏（pnl）",
    "pnl_pct": "浮动盈亏比例（pnl_pct）",
    "position_status": "持仓状态（position_status）",
    "position_qty": "模拟持仓数量（position_qty）",
    "avg_cost": "平均成本（avg_cost）",
    "unrealized_pnl": "浮动盈亏（unrealized_pnl）",
    "unrealized_pnl_pct": "浮动盈亏比例（unrealized_pnl_pct）",
    "holding_days": "持仓天数（holding_days）",
    "position_action": "持仓处理建议（position_action）",
    "position_reason": "持仓建议理由（position_reason）",
    "add_condition": "加仓条件（add_condition）",
    "reduce_condition": "减仓条件（reduce_condition）",
    "exit_condition": "退出条件（exit_condition）",
    "trigger_condition": "触发条件（trigger_condition）",
    "invalidation_condition": "失效条件（invalidation_condition）",
    "advice_reason": "建议理由（advice_reason）",
    "risk_status": "风险状态（risk_status）",
    "risk_status_cn": "风险状态说明（risk_status_cn）",
    "match_note": "匹配说明（match_note）",
    "note": "备注（note）",
    "risk_type": "风险类型（risk_type）",
    "risk_level": "风险等级（risk_level）",
    "detail": "说明（detail）",
    "suggested_action": "建议动作（suggested_action）",
    "table_name": "数据表（table_name）",
    "row_count": "行数（row_count）",
    "distinct_symbols": "股票数量（distinct_symbols）",
    "latest_date": "最新日期（latest_date）",
    "note": "说明（note）",
    "config_key": "配置项（config_key）",
    "config_value": "配置值（config_value）",
    "as_of_trade_date": "回看截止交易日（as_of_trade_date）",
    "sample_period": "样本区间（sample_period）",
    "horizons": "回看周期（horizons）",
    "candidate_sample_count": "候选样本数量（candidate_sample_count）",
    "valid_sample_count": "有效样本数量（valid_sample_count）",
    "insufficient_forward_data_count": "数据不足数量（insufficient_forward_data_count）",
    "total_score_group_summary": "综合分分组摘要（total_score_group_summary）",
    "elder_review_summary": "埃尔德复核摘要（elder_review_summary）",
    "entry_zone_summary": "买入区间摘要（entry_zone_summary）",
    "watchlist_summary": "观察池状态摘要（watchlist_summary）",
    "key_findings": "主要发现（key_findings）",
    "data_quality_summary": "数据质量提示（data_quality_summary）",
    "lookback_report_path": "完整回看报告路径（lookback_report_path）",
}


@dataclass(frozen=True)
class WorkbookExportResult:
    """Summary returned after exporting a daily research workbook."""

    output_path: Path
    trade_date: str
    strategy_rows: int
    elder_rows: int
    entry_zone_rows: int
    watchlist_rows: int
    external_position_rows: int
    simulated_advice_rows: int = 0
    lookback_summary_rows: int = 0


@dataclass(frozen=True)
class DailyResearchView:
    """User-visible daily research tables shared by Excel and Streamlit."""

    trade_date: str
    strategy_sheet: pd.DataFrame
    entry_sheet: pd.DataFrame
    entry_missing_sheet: pd.DataFrame
    watchlist_sheet: pd.DataFrame
    external_sheet: pd.DataFrame
    simulated_advice_sheet: pd.DataFrame
    lookback_sheet: pd.DataFrame
    include_lookback_summary: bool
    lookback_note: str


def export_daily_research_workbook(
    *,
    trade_date: str | None = None,
    output_path: str | Path | None = None,
    include_external_positions: bool = True,
    include_data_quality: bool = True,
    lookback_status_path: str | Path | None = None,
    settings: Settings | Any | None = None,
    store: DuckDBStore | None = None,
) -> WorkbookExportResult:
    """Export the latest local research state to a read-only Excel workbook.

    The export reads only existing DuckDB tables. It does not update market data,
    recompute factors, rerank candidates, or write back to DuckDB.
    """
    resolved_settings = settings or get_settings()
    resolved_store = store or DuckDBStore(getattr(resolved_settings, "duckdb_path", None))

    strategy = _read_table(resolved_store, "strategy_result")
    factor_scores = _read_table(resolved_store, "factor_scores")
    daily_price = _read_table(resolved_store, "daily_price")
    entry_zones = _read_table(resolved_store, "entry_zone_snapshots")
    watchlist = _read_table(resolved_store, "watchlist_daily_snapshots")
    external_positions = (
        _read_table(resolved_store, "external_position_snapshots")
        if include_external_positions
        else pd.DataFrame()
    )

    selected_trade_date = trade_date or _latest_trade_date(strategy, factor_scores)
    selected_trade_date = selected_trade_date or _latest_price_date(resolved_store) or ""

    entry_zones, entry_zone_auto_note = _ensure_entry_zones_for_trade_date(
        resolved_store,
        resolved_settings,
        entry_zones,
        selected_trade_date,
    )
    data_quality_snapshot = _safe_data_quality_snapshot(resolved_store, selected_trade_date)
    lookback_status = _read_lookback_status(lookback_status_path)
    backend_strategy_rows = len(_latest_by_date(strategy, "trade_date", selected_trade_date))
    view = build_daily_research_view_from_frames(
        strategy=strategy,
        entry_zones=entry_zones,
        watchlist=watchlist,
        external_positions=external_positions,
        daily_price=daily_price,
        trade_date=selected_trade_date,
        lookback_status=lookback_status,
    )
    embedded_elder_rows = _embedded_elder_row_count(view.strategy_sheet, view.watchlist_sheet)
    summary_sheet = _build_summary_sheet(
        selected_trade_date=selected_trade_date,
        output_path=output_path,
        strategy_rows=len(view.strategy_sheet),
        elder_rows=embedded_elder_rows,
        entry_zone_rows=len(view.entry_sheet),
        entry_zone_source_rows=len(view.entry_sheet) + len(view.entry_missing_sheet),
        entry_zone_missing_rows=len(view.entry_missing_sheet),
        watchlist_rows=len(view.watchlist_sheet),
        external_position_rows=len(view.external_sheet),
        simulated_advice=view.simulated_advice_sheet,
        backend_strategy_rows=backend_strategy_rows,
        lookback_status=lookback_status if view.include_lookback_summary else None,
        lookback_note=_join_notes(view.lookback_note, entry_zone_auto_note),
        data_quality_snapshot=data_quality_snapshot,
    )

    resolved_output = _resolve_output_path(output_path, selected_trade_date)
    resolved_output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "00_摘要": summary_sheet,
        "01_今日候选": view.strategy_sheet,
        "03_买入区间": view.entry_sheet,
        "04_观察池": _empty_if_needed(view.watchlist_sheet, "暂无观察池跟踪数据。"),
        "06_外部模拟持仓": _empty_if_needed(view.external_sheet, "暂无外部模拟持仓数据。"),
        "13_模拟交易建议": _empty_if_needed(view.simulated_advice_sheet, "暂无模拟交易建议。"),
    }
    ordered_sheet_names = [*SHEET_NAMES]
    if not view.entry_missing_sheet.empty:
        sheets["12_买入区间缺失说明"] = view.entry_missing_sheet
        ordered_sheet_names.append("12_买入区间缺失说明")
    if view.include_lookback_summary:
        sheets["11_自动回看摘要"] = view.lookback_sheet
        ordered_sheet_names.append("11_自动回看摘要")
    elif view.lookback_note:
        sheets["11_自动回看状态"] = _build_lookback_status_sheet(lookback_status, selected_trade_date, view.lookback_note)
        ordered_sheet_names.append("11_自动回看状态")
    for sheet_name in ordered_sheet_names:
        _write_sheet(workbook, sheet_name, sheets[sheet_name])
    workbook.save(resolved_output)

    return WorkbookExportResult(
        output_path=resolved_output,
        trade_date=selected_trade_date,
        strategy_rows=len(view.strategy_sheet),
        elder_rows=embedded_elder_rows,
        entry_zone_rows=len(view.entry_sheet),
        watchlist_rows=len(view.watchlist_sheet),
        external_position_rows=len(view.external_sheet),
        simulated_advice_rows=len(view.simulated_advice_sheet),
        lookback_summary_rows=len(view.lookback_sheet),
    )


def _read_table(store: DuckDBStore, table_name: str) -> pd.DataFrame:
    """Read a table through a read-only DuckDB connection, returning empty on failure."""
    try:
        return store.read_table(table_name, read_only=True)
    except DuckDBStoreError:
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _read_query(store: DuckDBStore, sql: str) -> pd.DataFrame:
    """Read a query through a read-only DuckDB connection, returning empty on failure."""
    try:
        with store.connect(read_only=True) as connection:
            return connection.execute(sql).fetchdf()
    except Exception:
        return pd.DataFrame()


def build_daily_research_view_from_frames(
    *,
    strategy: pd.DataFrame,
    entry_zones: pd.DataFrame,
    watchlist: pd.DataFrame,
    external_positions: pd.DataFrame | None = None,
    daily_price: pd.DataFrame | None = None,
    trade_date: str,
    lookback_status: dict[str, Any] | None = None,
) -> DailyResearchView:
    """Build the visible daily research tables from already-loaded frames."""
    strategy_sheet = _build_strategy_sheet(strategy, trade_date, daily_price)
    latest_watchlist_sheet = _latest_by_date(watchlist, "trade_date", trade_date)
    current_watchlist_sheet = _current_watchlist_scope(latest_watchlist_sheet)
    current_watchlist_sheet = _attach_elder_fields(current_watchlist_sheet, daily_price, "观察池")
    watchlist_sheet = _with_display_order(_preferred_columns(current_watchlist_sheet, _watchlist_columns()))
    entry_sheet = _build_entry_zone_sheet(entry_zones, strategy_sheet, watchlist_sheet, trade_date)
    entry_missing_sheet = _build_entry_zone_missing_sheet(entry_zones, strategy_sheet, watchlist_sheet, trade_date, entry_sheet)
    external_sheet = _latest_external_positions(external_positions if isinstance(external_positions, pd.DataFrame) else pd.DataFrame())
    simulated_advice_sheet = build_simulated_trading_advice(
        strategy=strategy_sheet,
        watchlist=watchlist_sheet,
        entry_zones=entry_sheet,
        entry_missing=entry_missing_sheet,
        external_positions=external_sheet,
        trade_date=trade_date,
    )
    include_lookback_summary, lookback_note = _lookback_summary_decision(lookback_status, trade_date)
    lookback_sheet = _build_lookback_summary_sheet(lookback_status) if include_lookback_summary else pd.DataFrame()
    return DailyResearchView(
        trade_date=trade_date,
        strategy_sheet=strategy_sheet,
        entry_sheet=entry_sheet,
        entry_missing_sheet=entry_missing_sheet,
        watchlist_sheet=watchlist_sheet,
        external_sheet=external_sheet,
        simulated_advice_sheet=simulated_advice_sheet,
        lookback_sheet=lookback_sheet,
        include_lookback_summary=include_lookback_summary,
        lookback_note=lookback_note,
    )


def _ensure_entry_zones_for_trade_date(
    store: DuckDBStore,
    settings: Settings | Any,
    entry_zones: pd.DataFrame,
    trade_date: str,
) -> tuple[pd.DataFrame, str]:
    """Use the existing entry-zone job to fill missing same-day snapshots."""
    if not trade_date or not _rows_at_date(entry_zones, "trade_date", trade_date).empty:
        return entry_zones, ""
    try:
        from core.jobs.calculate_entry_zones import calculate_entry_zones

        result = calculate_entry_zones(settings=settings, store=store, quiet=True)
    except Exception:
        return entry_zones, "买入区间：当日快照缺失，自动补算未完成；可运行 .venv/bin/python -m core.jobs.calculate_entry_zones。"
    refreshed = _read_table(store, "entry_zone_snapshots")
    result_date = _normalize_trade_date(result.get("trade_date") if isinstance(result, dict) else "")
    if result_date == _normalize_trade_date(trade_date) and not _rows_at_date(refreshed, "trade_date", trade_date).empty:
        return refreshed, "买入区间：导出前已补算当日买入区间。"
    return refreshed if not refreshed.empty else entry_zones, "买入区间：当日快照仍缺失；可运行 .venv/bin/python -m core.jobs.calculate_entry_zones 后重新导出。"


def _latest_trade_date(strategy: pd.DataFrame, factor_scores: pd.DataFrame) -> str:
    """Return latest persisted selection/factor date."""
    for frame in (strategy, factor_scores):
        if "trade_date" in frame.columns and not frame.empty:
            value = frame["trade_date"].dropna().astype(str).max()
            if value:
                return value
    return ""


def _latest_price_date(store: DuckDBStore) -> str:
    latest = _read_query(store, "SELECT MAX(trade_date) AS latest_trade_date FROM daily_price")
    if latest.empty:
        return ""
    value = latest.iloc[0].get("latest_trade_date")
    return "" if pd.isna(value) else str(value)


def _latest_by_date(frame: pd.DataFrame, date_column: str, preferred_date: str = "") -> pd.DataFrame:
    if frame.empty or date_column not in frame.columns:
        return pd.DataFrame()
    dates = frame[date_column].dropna().astype(str)
    if dates.empty:
        return pd.DataFrame()
    selected_date = preferred_date if preferred_date and preferred_date in set(dates) else dates.max()
    return frame[frame[date_column].astype(str) == selected_date].copy()


def _rows_at_date(frame: pd.DataFrame, date_column: str, trade_date: str) -> pd.DataFrame:
    if frame.empty or date_column not in frame.columns or not trade_date:
        return pd.DataFrame()
    normalized_target = _normalize_trade_date(trade_date)
    normalized_dates = frame[date_column].map(_normalize_trade_date)
    return frame[normalized_dates == normalized_target].copy()


def _normalize_trade_date(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y%m%d")
    text = str(value).strip()
    if not text:
        return ""
    return text.replace("-", "")[:8]


def _build_strategy_sheet(strategy: pd.DataFrame, trade_date: str, price_df: pd.DataFrame | None = None) -> pd.DataFrame:
    frame = _latest_by_date(strategy, "trade_date", trade_date)
    if frame.empty:
        return _message_frame("暂无本地选股结果。请先运行 python -m core.jobs.run_daily_workflow --doctor-before-run --skip-update --format all。")
    frame = frame.copy()
    if "rank" in frame.columns:
        frame["candidate_rank"] = pd.to_numeric(frame["rank"], errors="coerce")
        frame = frame.sort_values(["candidate_rank", "ts_code"], na_position="last")
    elif "total_score" in frame.columns:
        frame = frame.sort_values("total_score", ascending=False, na_position="last")
    if "ts_code" in frame.columns:
        frame = frame.drop_duplicates("ts_code", keep="first")
    frame = frame.head(SELECTION_DISPLAY_TOP_N)
    frame = _attach_elder_fields(frame, price_df, "今日候选")
    columns = [
        "display_order",
        "trade_date",
        "ts_code",
        "name",
        "industry",
        "list_date",
        "close",
        "pe",
        "pb",
        "total_score",
        "trend_score",
        "momentum_score",
        "liquidity_score",
        "fundamental_score",
        "volatility_score",
        "quality_score",
        "valuation_score",
        "risk_score",
        "select_reason",
        "risk_note",
        *ELDER_DISPLAY_COLUMNS,
    ]
    frame = _ensure_columns(frame, columns)
    return _with_display_order(_preferred_columns(frame, columns))


def _attach_elder_fields(frame: pd.DataFrame, price_df: pd.DataFrame | None, scope: str) -> pd.DataFrame:
    """Attach Elder display fields in memory without changing persisted tables."""
    if frame.empty or "message" in frame.columns:
        return frame.copy()
    result = frame.copy()
    for column in ELDER_DISPLAY_COLUMNS:
        if column not in result.columns:
            result[column] = pd.NA
    if price_df is None or price_df.empty or "ts_code" not in result.columns:
        return result
    candidates = result.copy()
    candidates["review_scope"] = scope
    try:
        review = build_elder_review(candidates, price_df)
    except Exception:
        return result
    if review.empty or "ts_code" not in review.columns:
        return result
    review = review.drop_duplicates("ts_code", keep="last").set_index(review["ts_code"].astype(str))
    codes = result["ts_code"].astype(str)
    for column in ELDER_DISPLAY_COLUMNS:
        if column not in review.columns:
            continue
        computed = codes.map(review[column])
        existing = result[column]
        missing = existing.isna() | existing.astype(str).str.strip().eq("")
        result.loc[missing, column] = computed[missing]
    return result


def _current_watchlist_scope(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or "watch_status" not in frame.columns:
        return frame.copy().head(WATCHLIST_DISPLAY_TOP_N)
    current_statuses = {"active", "entry_zone", "triggered", "active_watch", "strong_watch", "wait_pullback", "near_buy_zone"}
    result = frame[frame["watch_status"].fillna("active").astype(str).isin(current_statuses)].copy()
    if "ts_code" in result.columns:
        result = result.drop_duplicates("ts_code", keep="last")
    return result.reset_index(drop=True).head(WATCHLIST_DISPLAY_TOP_N)


def _build_entry_zone_sheet(
    entry_zones: pd.DataFrame,
    strategy: pd.DataFrame,
    watchlist: pd.DataFrame,
    trade_date: str,
) -> pd.DataFrame:
    columns = [
        "display_order",
        "source",
        "trade_date",
        "ts_code",
        "name",
        "close",
        "ema13",
        "ema22",
        "ema60",
        "support_20d",
        "support_60d",
        "resistance_20d",
        "resistance_60d",
        "nearest_support",
        "nearest_resistance",
        "atr_14",
        "entry_low",
        "entry_high",
        "entry_mid",
        "stop_loss",
        "target_price",
        "risk_pct",
        "reward_pct",
        "reward_risk_ratio",
        "entry_zone_status",
        "entry_zone_status_cn",
        "chase_risk",
        "chase_risk_cn",
        "price_action_note",
        "entry_reason",
        "risk_note",
    ]
    frame = _rows_at_date(entry_zones, "trade_date", trade_date)
    if frame.empty:
        return pd.DataFrame(columns=columns)
    if "source" not in frame.columns:
        frame["source"] = ""
    if "source" in frame.columns:
        frame = frame[frame["source"].fillna("").astype(str).isin({"selection", "watchlist", ""})].copy()
    source_scope = _entry_zone_source_scope(strategy, watchlist)
    source_map = dict(zip(source_scope.get("ts_code", pd.Series(dtype=str)).astype(str), source_scope.get("source", pd.Series(dtype=str)).astype(str)))
    visible_codes = set(source_map)
    if visible_codes and "ts_code" in frame.columns:
        frame = frame[frame["ts_code"].astype(str).isin(visible_codes)].copy()
    if not frame.empty and "ts_code" in frame.columns:
        frame["source"] = frame["ts_code"].astype(str).map(lambda code: source_map.get(code, ""))
        source_priority = {"selection": 0, "": 1, "watchlist": 2}
        frame["_source_priority"] = frame["source"].fillna("").astype(str).map(source_priority).fillna(3)
        frame = frame.sort_values(["_source_priority", "ts_code"]).drop_duplicates("ts_code", keep="first")
        frame = frame.drop(columns=["_source_priority"])
    return _with_display_order(_preferred_columns(frame, columns))


def _build_entry_zone_missing_sheet(
    entry_zones: pd.DataFrame,
    strategy: pd.DataFrame,
    watchlist: pd.DataFrame,
    trade_date: str,
    entry_sheet: pd.DataFrame,
) -> pd.DataFrame:
    """Build missing entry-zone explanation for the visible research universe."""
    source_scope = _entry_zone_source_scope(strategy, watchlist)
    columns = ["display_order", "trade_date", "ts_code", "name", "source", "missing_reason"]
    if source_scope.empty:
        return pd.DataFrame(columns=columns)
    present_codes = _frame_codes(entry_sheet)
    missing = source_scope[~source_scope["ts_code"].astype(str).isin(present_codes)].copy()
    if missing.empty:
        return pd.DataFrame(columns=columns)
    same_day = _rows_at_date(entry_zones, "trade_date", trade_date)
    same_day_codes = _frame_codes(same_day)
    missing["trade_date"] = trade_date
    missing["missing_reason"] = missing["ts_code"].astype(str).map(
        lambda code: "缺少买入区间数据，请重新运行 calculate_entry_zones" if code in same_day_codes else "未生成买入区间快照"
    )
    return _with_display_order(_preferred_columns(missing, columns))


def _entry_zone_source_scope(strategy: pd.DataFrame, watchlist: pd.DataFrame) -> pd.DataFrame:
    """Return Top10 selection + Top30 watchlist visible universe with selection priority."""
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for source, frame in [("selection", strategy), ("watchlist", watchlist)]:
        if not isinstance(frame, pd.DataFrame) or frame.empty or "ts_code" not in frame.columns:
            continue
        for _, row in frame.iterrows():
            code = str(row.get("ts_code") or "").strip()
            if not code or code in seen:
                continue
            seen.add(code)
            rows.append({"ts_code": code, "name": row.get("name"), "source": source})
    return pd.DataFrame(rows, columns=["ts_code", "name", "source"])


def _visible_research_codes(strategy: pd.DataFrame, watchlist: pd.DataFrame) -> set[str]:
    return _frame_codes(_entry_zone_source_scope(strategy, watchlist))


def _frame_codes(frame: pd.DataFrame) -> set[str]:
    codes: list[str] = []
    if frame.empty or "message" in frame.columns or "ts_code" not in frame.columns:
        return set()
    codes.extend(frame["ts_code"].dropna().astype(str).tolist())
    return set(dict.fromkeys(code for code in codes if code))


def _watchlist_columns() -> list[str]:
    return [
        "display_order",
        "trade_date",
        "ts_code",
        "name",
        "current_close",
        "pe",
        "pb",
        "total_score",
        "total_score_change",
        "is_top_n",
        "is_new_candidate",
        "first_selected_date",
        "last_selected_date",
        "selected_count_5d",
        "selected_count_10d",
        "consecutive_selected_days",
        "watch_status",
        "watch_status_label",
        "watch_days",
        "entry_reason",
        "watch_reason",
        "elder_score",
        "action_hint",
        "elder_reason",
        "weekly_trend",
        "daily_pullback",
        "force_signal",
        "elder_ray_signal",
        "daily_note",
    ]


def _latest_external_positions(external_positions: pd.DataFrame) -> pd.DataFrame:
    frame = _latest_by_date(external_positions, "snapshot_date")
    columns = [
        "display_order",
        "snapshot_date",
        "platform",
        "account_name",
        "ts_code",
        "name",
        "quantity",
        "cost_price",
        "current_price",
        "market_value",
        "pnl",
        "pnl_pct",
        "stop_loss",
        "target_price",
        "entry_low",
        "entry_high",
        "reward_risk_ratio",
        "position_status",
        "risk_status",
        "risk_status_cn",
        "match_note",
        "note",
    ]
    return _with_display_order(_preferred_columns(frame, columns))


def _build_risk_sheet(
    entry_zones: pd.DataFrame,
    watchlist: pd.DataFrame,
    external_positions: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if not entry_zones.empty and "message" not in entry_zones.columns:
        for _, row in entry_zones.iterrows():
            risk_items: list[str] = []
            if str(row.get("chase_risk", "")).lower() == "high":
                risk_items.append("追高风险高")
            if str(row.get("entry_zone_status", "")).lower() in {"weak_no_entry", "insufficient_data"}:
                risk_items.append(str(row.get("entry_zone_status_cn") or row.get("entry_zone_status")))
            if pd.notna(row.get("reward_risk_ratio")) and float(row.get("reward_risk_ratio")) < 2:
                risk_items.append("盈亏比低于 2")
            if risk_items:
                rows.append(
                    {
                        "source": "买入区间",
                        "ts_code": row.get("ts_code"),
                        "name": row.get("name"),
                        "risk_type": "；".join(risk_items),
                        "detail": row.get("risk_note") or row.get("price_action_note"),
                        "suggested_action": "人工复核",
                    }
                )
    if not watchlist.empty and "message" not in watchlist.columns:
        for _, row in watchlist.iterrows():
            status = str(row.get("watch_status") or "")
            if status in {"overheated", "weakening", "invalidated"}:
                rows.append(
                    {
                        "source": "观察池",
                        "ts_code": row.get("ts_code"),
                        "name": row.get("name"),
                        "risk_type": row.get("watch_status_label") or status,
                        "detail": row.get("daily_note") or row.get("elder_reason"),
                        "suggested_action": "人工复核",
                    }
                )
    if not external_positions.empty and "message" not in external_positions.columns:
        for _, row in external_positions.iterrows():
            risk_status = str(row.get("risk_status") or "")
            if risk_status and risk_status not in {"normal", "matched", "ok"}:
                rows.append(
                    {
                        "source": "外部模拟持仓",
                        "ts_code": row.get("ts_code"),
                        "name": row.get("name"),
                        "risk_type": row.get("risk_status_cn") or risk_status,
                        "detail": row.get("match_note") or row.get("note"),
                        "suggested_action": "人工复核",
                    }
                )
    return pd.DataFrame(rows)


def _build_data_quality_sheet(store: DuckDBStore, trade_date: str, snapshot: dict[str, Any] | None = None) -> pd.DataFrame:
    snapshot_rows = _data_quality_snapshot_rows(snapshot or {})
    metrics = [
        _table_metric(store, "stock_basic", "ts_code"),
        _table_metric(store, "daily_price", "ts_code", "trade_date"),
        _table_metric(store, "daily_basic", "ts_code", "trade_date"),
        _table_metric(store, "factor_scores", "ts_code", "trade_date"),
        _table_metric(store, "strategy_result", "ts_code", "trade_date"),
        _table_metric(store, "entry_zone_snapshots", "ts_code", "trade_date"),
        _table_metric(store, "watchlist_daily_snapshots", "ts_code", "trade_date"),
        _table_metric(store, "external_position_snapshots", "ts_code", "snapshot_date"),
    ]
    quality = pd.DataFrame(metrics)
    if snapshot_rows:
        quality = pd.concat([pd.DataFrame(snapshot_rows), quality], ignore_index=True, sort=False)
    if trade_date:
        quality.loc[len(quality)] = {
            "table_name": "export_scope",
            "row_count": None,
            "distinct_symbols": None,
            "latest_date": trade_date,
            "note": "工作簿默认使用该日期的本地持久化结果。",
        }
    missing_metrics = _missing_value_metrics(store, trade_date)
    if missing_metrics:
        quality = pd.concat([quality, pd.DataFrame(missing_metrics)], ignore_index=True, sort=False)
    return quality


def _data_quality_snapshot_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    if not snapshot:
        return []
    return [
        {"table_name": "data_quality_snapshot", "row_count": "", "distinct_symbols": "", "latest_date": snapshot.get("latest_completed_trade_date"), "note": f"data_quality_status={snapshot.get('data_quality_status')}"},
        {"table_name": "latest_daily_price", "row_count": "", "distinct_symbols": snapshot.get("latest_daily_price_symbol_count", 0), "latest_date": snapshot.get("latest_completed_trade_date"), "note": f"coverage={float(snapshot.get('latest_daily_price_coverage_rate', 0.0) or 0.0):.2%}"},
        {"table_name": "latest_daily_basic", "row_count": "", "distinct_symbols": snapshot.get("latest_daily_basic_symbol_count", 0), "latest_date": snapshot.get("latest_completed_trade_date"), "note": f"coverage={float(snapshot.get('latest_daily_basic_coverage_rate', 0.0) or 0.0):.2%}"},
        {"table_name": "latest_adj_factor", "row_count": "", "distinct_symbols": snapshot.get("latest_adj_factor_symbol_count", 0), "latest_date": snapshot.get("latest_completed_trade_date"), "note": f"coverage={float(snapshot.get('latest_adj_factor_coverage_rate', 0.0) or 0.0):.2%}"},
        {"table_name": "any_daily_price", "row_count": "", "distinct_symbols": snapshot.get("any_daily_price_symbol_count", 0), "latest_date": "", "note": f"coverage={float(snapshot.get('any_daily_price_coverage_rate', 0.0) or 0.0):.2%}"},
        {"table_name": "history_missing", "row_count": "", "distinct_symbols": snapshot.get("history_missing_symbol_count", 0), "latest_date": "", "note": "完全缺行情股票数量"},
    ]


def _table_metric(
    store: DuckDBStore,
    table_name: str,
    symbol_column: str | None = None,
    date_column: str | None = None,
) -> dict[str, Any]:
    symbol_expr = f"COUNT(DISTINCT {symbol_column}) AS distinct_symbols" if symbol_column else "NULL AS distinct_symbols"
    date_expr = f"MAX({date_column}) AS latest_date" if date_column else "NULL AS latest_date"
    frame = _read_query(
        store,
        f"SELECT COUNT(*) AS row_count, {symbol_expr}, {date_expr} FROM {table_name}",
    )
    if frame.empty:
        return {
            "table_name": table_name,
            "row_count": 0,
            "distinct_symbols": 0,
            "latest_date": "",
            "note": "表不存在或不可读。",
        }
    row = frame.iloc[0]
    return {
        "table_name": table_name,
        "row_count": _blank_if_na(row.get("row_count")),
        "distinct_symbols": _blank_if_na(row.get("distinct_symbols")),
        "latest_date": _blank_if_na(row.get("latest_date")),
        "note": "",
    }


def _missing_value_metrics(store: DuckDBStore, trade_date: str) -> list[dict[str, Any]]:
    """Return missing-value metrics with wording that distinguishes columns from values."""
    metrics: list[dict[str, Any]] = []
    targets = [
        ("factor_scores", "fundamental_score", "基本面分（fundamental_score）"),
        ("strategy_result", "fundamental_score", "基本面分（fundamental_score）"),
        ("strategy_result", "pe", "市盈率（pe）"),
        ("strategy_result", "pb", "市净率（pb）"),
    ]
    for table_name, column_name, label in targets:
        if not _table_has_column(store, table_name, column_name):
            metrics.append(
                {
                    "table_name": table_name,
                    "row_count": "",
                    "distinct_symbols": "",
                    "latest_date": trade_date,
                    "note": f"字段不存在：{label}",
                }
            )
            continue
        where = f"WHERE trade_date = '{trade_date}'" if trade_date and _table_has_column(store, table_name, "trade_date") else ""
        frame = _read_query(
            store,
            f"SELECT COUNT(*) AS row_count, SUM(CASE WHEN {column_name} IS NULL THEN 1 ELSE 0 END) AS missing_count FROM {table_name} {where}",
        )
        if frame.empty:
            continue
        row = frame.iloc[0]
        raw_missing = row.get("missing_count")
        missing_count = 0 if pd.isna(raw_missing) else int(raw_missing or 0)
        if missing_count > 0:
            note_prefix = FUNDAMENTAL_SCORE_MISSING_NOTE_PREFIX if column_name == "fundamental_score" else f"{label}缺失记录数"
            metrics.append(
                {
                    "table_name": table_name,
                    "row_count": _blank_if_na(row.get("row_count")),
                    "distinct_symbols": "",
                    "latest_date": trade_date,
                    "note": f"{note_prefix}：{missing_count}",
                }
            )
    return metrics


def _table_has_column(store: DuckDBStore, table_name: str, column_name: str) -> bool:
    frame = _read_query(store, f"DESCRIBE {table_name}")
    if frame.empty or "column_name" not in frame.columns:
        return False
    return column_name in set(frame["column_name"].astype(str))


def _settings_sheet(settings: Any) -> pd.DataFrame:
    if hasattr(settings, "model_dump"):
        values = settings.model_dump()
    else:
        values = {key: value for key, value in vars(settings).items() if not key.startswith("_")}
    rows = []
    for key, value in sorted(values.items()):
        lower = key.lower()
        value_text = str(value)
        if any(marker in lower for marker in SENSITIVE_KEYWORDS):
            continue
        if any(marker in value_text.lower() for marker in SENSITIVE_VALUE_MARKERS):
            continue
        rows.append({"config_key": key, "config_value": value_text})
    return pd.DataFrame(rows)


def _build_summary_sheet(
    *,
    selected_trade_date: str,
    output_path: str | Path | None,
    strategy_rows: int,
    elder_rows: int,
    entry_zone_rows: int,
    watchlist_rows: int,
    external_position_rows: int,
    entry_zone_source_rows: int = 0,
    entry_zone_missing_rows: int = 0,
    backend_strategy_rows: int | None = None,
    simulated_advice: pd.DataFrame | None = None,
    lookback_status: dict[str, Any] | None = None,
    lookback_note: str = "",
    data_quality_snapshot: dict[str, Any] | None = None,
) -> pd.DataFrame:
    advice_counts = summarize_simulated_trading_advice(simulated_advice)
    rows = [
        {"metric": "导出时间", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"metric": "研究日期", "value": selected_trade_date or "暂无"},
        {"metric": "后端候选池数量", "value": backend_strategy_rows if backend_strategy_rows is not None else strategy_rows},
        {"metric": "今日候选数量", "value": strategy_rows},
        {"metric": "日报展示候选数量", "value": strategy_rows},
        {"metric": "已嵌入埃尔德字段记录数量", "value": elder_rows},
        {"metric": "买入区间记录数量", "value": entry_zone_rows},
        {"metric": "买入区间来源展示全集数量", "value": entry_zone_source_rows},
        {"metric": "买入区间缺失数量", "value": entry_zone_missing_rows},
        {"metric": "观察池记录数量", "value": watchlist_rows},
        {"metric": "观察池展示数量", "value": watchlist_rows},
        {"metric": "外部模拟持仓记录数量", "value": external_position_rows},
        {"metric": "模拟交易建议数量", "value": advice_counts["total"]},
        {"metric": "可模拟买入数量", "value": advice_counts["buy"]},
        {"metric": "等待回调数量", "value": advice_counts["wait_pullback"]},
        {"metric": "继续观察数量", "value": advice_counts["observe"]},
        {"metric": "暂缓数量", "value": advice_counts["pause"]},
        {"metric": "剔除数量", "value": advice_counts["remove"]},
        {"metric": "模拟持仓跟踪数量", "value": advice_counts["holding"]},
        {"metric": "建议继续持有数量", "value": advice_counts["hold"]},
        {"metric": "建议加仓数量", "value": advice_counts["add"]},
        {"metric": "建议减仓数量", "value": advice_counts["reduce"]},
        {"metric": "建议卖出数量", "value": advice_counts["sell"]},
    ]
    if data_quality_snapshot:
        rows.extend(
            [
                {"metric": "数据质量状态", "value": data_quality_snapshot.get("data_quality_status") or "unknown"},
                {"metric": "正式全市场研究结果可用", "value": "是" if data_quality_snapshot.get("formal_result_usable") is True else "否"},
                {"metric": "最新交易日 daily_price 覆盖", "value": f"{data_quality_snapshot.get('latest_daily_price_symbol_count', 0)} / {data_quality_snapshot.get('configured_symbol_count', 0)}"},
                {"metric": "任意历史行情覆盖", "value": f"{data_quality_snapshot.get('any_daily_price_symbol_count', 0)} / {data_quality_snapshot.get('configured_symbol_count', 0)}"},
            ]
        )
        if data_quality_snapshot.get("data_quality_status") == "poor":
            rows.append({"metric": "数据质量提示", "value": "当前结果仅供流程检查，不代表完整全市场筛选。"})
    if lookback_status:
        rows.extend(
            [
                {"metric": "最近一次自动回看状态", "value": lookback_status.get("status") or "暂无"},
                {"metric": "回看截止交易日", "value": lookback_status.get("as_of_trade_date") or "暂无"},
                {"metric": "回看有效样本数量", "value": lookback_status.get("valid_sample_count", 0)},
                {"metric": "完整回看报告路径", "value": lookback_status.get("generated_report_path") or "暂无"},
            ]
        )
    else:
        rows.append({"metric": "最近一次自动回看状态", "value": lookback_note or "尚无自动回看记录。"})
    rows.extend(
        [
            {"metric": "输出文件", "value": str(output_path or "默认 reports/daily_research_*.xlsx")},
            {"metric": "说明", "value": "仅供个人研究使用，不自动交易。"},
        ]
    )
    return pd.DataFrame(rows)


def _safe_data_quality_snapshot(store: DuckDBStore, trade_date: str) -> dict[str, Any]:
    try:
        return build_data_quality_snapshot(db_path=store.db_path, research_trade_date=trade_date, latest_completed_trade_date=trade_date)
    except Exception:
        return {
            "data_quality_status": "unknown",
            "formal_result_usable": False,
            "formal_result_warning_reason": "数据质量快照不可用。",
            "latest_completed_trade_date": trade_date,
        }


def _read_lookback_status(status_path: str | Path | None = None) -> dict[str, Any] | None:
    """Read the latest lookback status JSON without requiring it to exist."""
    path = Path(status_path) if status_path else PROJECT_ROOT / "data" / "runtime" / "lookback_analysis_status.json"
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {"status": "failed", "summary": "自动回看状态文件不可读。"}
    return payload if isinstance(payload, dict) else None


def _lookback_summary_decision(status: dict[str, Any] | None, trade_date: str) -> tuple[bool, str]:
    """Decide whether a lookback status should be expanded in this workbook."""
    if not status:
        return False, "尚无自动回看记录。"
    status_date = _normalize_trade_date(status.get("as_of_trade_date") or status.get("end_date") or "")
    workbook_date = _normalize_trade_date(trade_date)
    if workbook_date and status_date and status_date < workbook_date:
        return False, f"自动回看：最近回看截止 {status_date}，当前研究日期 {workbook_date} 需要刷新当日回看；本日报不把旧回看冒充当日结果。"
    valid_count = _safe_int(status.get("valid_sample_count"))
    candidate_count = _safe_int(status.get("candidate_sample_count"))
    if valid_count <= 0 or candidate_count <= 0:
        return False, "自动回看：最近回看有效样本不足，本日报未展开回看摘要。"
    if _lookback_summary_is_uninformative(status):
        return False, "自动回看：最近回看暂无可统计样本，本日报未展开回看摘要。"
    return True, ""


def _build_lookback_status_sheet(status: dict[str, Any] | None, trade_date: str, note: str) -> pd.DataFrame:
    display = build_lookback_status_display(status, trade_date)
    status_date = str(display["summary"].get("回看截止交易日") or "")
    return pd.DataFrame(
        [
            {"metric": "当前研究日期", "value": str(display["summary"].get("当前研究日期") or _normalize_trade_date(trade_date) or "暂无")},
            {"metric": "最近回看截止日期", "value": status_date or "暂无"},
            {"metric": "最近回看有效样本数", "value": (status or {}).get("valid_sample_count", 0) if status else 0},
            {"metric": "最近回看数据不足数", "value": (status or {}).get("insufficient_forward_data_count", 0) if status else 0},
            {"metric": "状态", "value": note or "需要刷新当日回看。"},
            {"metric": "建议命令", "value": str(display["summary"].get("建议命令") or f".venv/bin/python -m core.jobs.run_lookback_analysis --as-of {_normalize_trade_date(trade_date) or 'YYYYMMDD'} --format text")},
        ]
    )


def build_lookback_status_display(status: dict[str, Any] | None, current_research_trade_date: str) -> dict[str, Any]:
    """Return shared user-facing lookback status semantics for Excel and Streamlit."""
    status = status or {}
    lookback_date = _normalize_trade_date(status.get("as_of_trade_date") or status.get("end_date") or "")
    current_date = _normalize_trade_date(current_research_trade_date)
    is_current = bool(current_date and lookback_date and lookback_date == current_date)
    report_path_text = str(status.get("generated_report_path") or "")
    report_exists = bool(report_path_text and Path(report_path_text).exists())
    summary: dict[str, Any] = {
        "回看状态": "当前有效回看" if is_current else "需要刷新当日回看",
        "当前研究日期": current_date or "暂无",
        "回看截止交易日": lookback_date or "暂无",
        "样本区间": f"{status.get('start_date') or '暂无'} - {status.get('end_date') or '暂无'}",
        "回看周期": ",".join(str(item) for item in status.get("horizons", [])) if isinstance(status.get("horizons"), list) else status.get("horizons"),
        "候选样本数量": status.get("candidate_sample_count", 0),
        "有效样本数量": status.get("valid_sample_count", 0),
        "数据不足数量": status.get("insufficient_forward_data_count", 0),
        "主要发现": status.get("key_findings") or "暂无",
        "数据质量提示": status.get("data_quality_summary") or "暂无",
        "报告路径": report_path_text or "暂无",
        "报告文件状态": "存在" if report_exists else ("文件不存在，可能已清理" if report_path_text else "暂无"),
    }
    if not is_current and lookback_date:
        summary["说明"] = f"最近回看截止 {lookback_date}，当前研究日期 {current_date or '暂无'} 需要刷新当日回看。"
        summary["建议命令"] = f".venv/bin/python -m core.jobs.run_lookback_analysis --as-of {current_date or 'YYYYMMDD'} --format text"
    return {"is_current": is_current, "summary": summary, "report_exists": report_exists}


def _lookback_summary_is_uninformative(status: dict[str, Any]) -> bool:
    fields = [
        "total_score_group_summary",
        "elder_review_summary",
        "entry_zone_summary",
        "watchlist_summary",
        "key_findings",
    ]
    values = [str(status.get(field) or "").strip() for field in fields]
    non_empty = [value for value in values if value]
    if not non_empty:
        return True
    return all(("暂无可统计样本" in value or "无样本" in value) for value in non_empty)


def _safe_int(value: Any) -> int:
    try:
        if pd.isna(value):
            return 0
        return int(value)
    except (TypeError, ValueError):
        return 0


def _join_notes(*notes: str) -> str:
    return "；".join(note for note in notes if note)


def _build_lookback_summary_sheet(status: dict[str, Any] | None) -> pd.DataFrame:
    """Build a lightweight lookback summary sheet; never embed full detail rows."""
    if not status:
        return _message_frame("尚无自动回看记录。")
    sample_period = f"{status.get('start_date') or '暂无'} - {status.get('end_date') or '暂无'}"
    horizons = status.get("horizons", [])
    if isinstance(horizons, list):
        horizon_text = ",".join(str(item) for item in horizons)
    else:
        horizon_text = str(horizons or "")
    return pd.DataFrame(
        [
            {
                "as_of_trade_date": status.get("as_of_trade_date") or "",
                "sample_period": sample_period,
                "horizons": horizon_text,
                "candidate_sample_count": status.get("candidate_sample_count", 0),
                "valid_sample_count": status.get("valid_sample_count", 0),
                "insufficient_forward_data_count": status.get("insufficient_forward_data_count", 0),
                "total_score_group_summary": status.get("total_score_group_summary") or "",
                "elder_review_summary": status.get("elder_review_summary") or "",
                "entry_zone_summary": status.get("entry_zone_summary") or "",
                "watchlist_summary": status.get("watchlist_summary") or "",
                "key_findings": status.get("key_findings") or "",
                "data_quality_summary": status.get("data_quality_summary") or "",
                "lookback_report_path": status.get("generated_report_path") or "",
            }
        ]
    )


def _help_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"section": "工作簿用途", "description": "汇总今日候选、买入区间、当前观察池、外部模拟持仓和必要摘要。"},
            {"section": "排序口径", "description": "序号只代表当前 Sheet 当前显示顺序，不代表买入优先级。"},
            {"section": "rank 字段", "description": "默认不导出 rank / 排名字段。系统已提供综合分、各因子分、埃尔德分、买入区间、风险状态等字段，用户可自行筛选和排序。"},
            {"section": "字段命名", "description": "字段命名采用“中文名称（英文名）”格式。中文名称用于理解含义，英文名用于和代码、数据库字段或导出字段对应。"},
            {"section": "观察池", "description": "04_观察池是当前观察名单，展示当前仍需跟踪的股票，不代表买入清单。"},
            {"section": "观察池历史", "description": "观察池历史、退出、归档和失效记录保留在后台表中，每日研究工作簿默认只展示当前观察池。"},
            {"section": "埃尔德复核", "description": "Elder 复核已作为今日候选和当前观察池的附加判断字段展示，不再单独输出股票明细。"},
            {"section": "买入区间", "description": "买入区间、止损价、目标价、盈亏比是研究计划参考，不是交易指令。"},
            {"section": "只读原则", "description": "导出命令只读取 DuckDB，不更新行情、不重算因子、不改变 total_score。"},
            {"section": "缺失数据", "description": "某些工作表显示暂无数据时，请先运行对应本地命令生成持久化结果。"},
            {"section": "提示", "description": "个人研究工具，结果需自行复核。"},
        ]
    )


def _preferred_columns(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=columns)
    available = [column for column in columns if column in frame.columns]
    return frame.loc[:, available].copy()


def _ensure_columns(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    result = frame.copy()
    for column in columns:
        if column not in result.columns:
            result[column] = pd.NA
    return result


def _embedded_elder_row_count(strategy: pd.DataFrame, watchlist: pd.DataFrame) -> int:
    total = 0
    for frame in [strategy, watchlist]:
        if frame.empty or "message" in frame.columns or "elder_score" not in frame.columns:
            continue
        total += int(pd.to_numeric(frame["elder_score"], errors="coerce").notna().sum())
    return total


def _with_display_order(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    frame = frame.reset_index(drop=True).copy()
    if "display_order" in frame.columns:
        frame = frame.drop(columns=["display_order"])
    frame.insert(0, "display_order", range(1, len(frame) + 1))
    return frame


def _empty_if_needed(frame: pd.DataFrame, message: str) -> pd.DataFrame:
    if frame.empty:
        return _message_frame(message)
    return frame


def _message_frame(message: str) -> pd.DataFrame:
    return pd.DataFrame([{"message": message}])


def _resolve_output_path(output_path: str | Path | None, trade_date: str) -> Path:
    if output_path is not None:
        return Path(output_path)
    timestamp = datetime.now().strftime("%H%M%S")
    date_part = trade_date or datetime.now().strftime("%Y%m%d")
    return PROJECT_ROOT / "reports" / f"daily_research_{date_part}_{timestamp}.xlsx"


def _write_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    frame = _prepare_for_excel(frame)
    if frame.empty and len(frame.columns) == 0:
        frame = _prepare_for_excel(_message_frame("暂无数据。"))
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False):
        worksheet.append([_excel_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 36)


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, float):
        return round(value, 6)
    return value


def _prepare_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    """Apply user-facing workbook rules without changing persisted data."""
    frame = frame.copy()
    frame = _drop_rank_columns(frame)
    frame = frame.rename(columns={column: COLUMN_LABELS.get(column, column) for column in frame.columns})
    frame = frame.where(pd.notna(frame), None)
    return frame


def _drop_rank_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    rank_like = {column for column in frame.columns if _is_rank_column(column)}
    if not rank_like:
        return frame
    return frame.drop(columns=sorted(rank_like))


def _is_rank_column(column: str) -> bool:
    normalized = str(column).strip().lower()
    if normalized in RANK_COLUMNS:
        return True
    forbidden_labels = (
        "原始选股排名",
        "当日入选排名",
        "上次入选排名",
        "排名变化",
        "历史最佳入选排名",
        "最近入选排名",
    )
    return any(label in str(column) for label in forbidden_labels)


def _blank_if_na(value: Any) -> Any:
    if pd.isna(value):
        return ""
    return value


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Export daily research workbook.")
    parser.add_argument("--trade-date", default=None, help="Trade date in YYYYMMDD; default latest local selection date.")
    parser.add_argument("--output", default=None, help="Output .xlsx path; default reports/daily_research_*.xlsx.")
    parser.add_argument("--format", default="xlsx", choices=["xlsx"], help="Output format. Only xlsx is supported.")
    parser.add_argument("--lookback-status-path", default=None, help="Optional lookback status JSON path.")
    parser.add_argument(
        "--include-external-positions",
        dest="include_external_positions",
        action="store_true",
        default=True,
        help="Include external simulated positions sheet. Enabled by default.",
    )
    parser.add_argument(
        "--no-include-external-positions",
        dest="include_external_positions",
        action="store_false",
        help="Skip external simulated positions sheet.",
    )
    parser.add_argument(
        "--include-data-quality",
        dest="include_data_quality",
        action="store_true",
        default=True,
        help="Include data quality sheet. Enabled by default.",
    )
    parser.add_argument(
        "--no-include-data-quality",
        dest="include_data_quality",
        action="store_false",
        help="Skip data quality sheet.",
    )
    args = parser.parse_args(argv)

    result = export_daily_research_workbook(
        trade_date=args.trade_date,
        output_path=args.output,
        include_external_positions=args.include_external_positions,
        include_data_quality=args.include_data_quality,
        lookback_status_path=args.lookback_status_path,
    )
    print("每日研究工作簿导出完成")
    print(f"研究日期: {result.trade_date or '暂无'}")
    print(f"今日候选: {result.strategy_rows}")
    print(f"Elder 字段补充: {result.elder_rows}")
    print(f"买入区间: {result.entry_zone_rows}")
    print(f"观察池: {result.watchlist_rows}")
    print(f"外部模拟持仓: {result.external_position_rows}")
    print(f"模拟交易建议: {result.simulated_advice_rows}")
    print(f"自动回看摘要: {result.lookback_summary_rows}")
    print(f"输出文件: {result.output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
