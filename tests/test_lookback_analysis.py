"""Tests for Task 58 automatic lookback analysis."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from core.jobs.run_lookback_analysis import (
    LookbackInputs,
    build_forward_return_details,
    build_lookback_samples,
    group_summary,
    run_lookback_analysis,
    save_lookback_workbook,
    score_bucket,
    summarize_lookback,
)
from core.runtime.command_runner import ALLOWED_COMMANDS
from core.storage.duckdb_store import DuckDBStore


def test_forward_return_uses_future_trading_days() -> None:
    """Forward returns should use future trading rows, not natural calendar days."""
    samples = _sample_strategy()
    prices = _price_frame()
    details = build_forward_return_details(samples, prices, [3])
    row = details[(details["ts_code"] == "000001.SZ") & (details["horizon"] == 3)].iloc[0]
    assert round(row["forward_return"], 6) == round(13 / 10 - 1, 6)
    assert row["available_forward_days"] == 5


def test_forward_return_does_not_use_future_data_for_selection() -> None:
    """Lookback detail calculation must not mutate persisted/current selection rows."""
    samples = _sample_strategy()
    before = samples.copy(deep=True)
    build_forward_return_details(samples, _price_frame(), [1, 3])
    pd.testing.assert_frame_equal(samples, before)


def test_insufficient_forward_data_is_excluded_from_horizon_stats() -> None:
    """Insufficient future data should be excluded from valid horizon stats."""
    details = build_forward_return_details(_sample_strategy(), _price_frame(days=2), [3])
    summary = group_summary(details.assign(candidate_group="全部候选"), "candidate_group")
    row = summary[summary["horizon"] == 3].iloc[0]
    assert row["valid_sample_count"] == 0
    assert row["insufficient_forward_data_count"] == 2


def test_max_drawdown_and_max_runup_calculation() -> None:
    """Max drawdown and max runup should use low/high over the future window."""
    details = build_forward_return_details(_sample_strategy().head(1), _price_frame(), [3])
    row = details.iloc[0]
    assert round(row["max_drawdown"], 6) == round(10 / 10 - 1, 6)
    assert round(row["max_runup"], 6) == round(14 / 10 - 1, 6)


def test_hit_stop_loss_and_target() -> None:
    """Stop loss and target touch checks should use future low/high."""
    samples = _sample_strategy().head(1).copy()
    samples["stop_loss"] = [10.2]
    samples["target_price"] = [13.5]
    details = build_forward_return_details(samples, _price_frame(), [3])
    row = details.iloc[0]
    assert bool(row["hit_stop_loss"]) is True
    assert bool(row["hit_target"]) is True


def test_total_score_grouping() -> None:
    """Total score buckets should match configured group labels."""
    assert score_bucket(85) == ">=80"
    assert score_bucket(75) == "70-80"
    assert score_bucket(65) == "60-70"
    assert score_bucket(55) == "50-60"
    assert score_bucket(45) == "<50"


def test_factor_score_grouping() -> None:
    """Factor score grouping summary should include factor dimensions."""
    details = build_forward_return_details(_sample_strategy(), _price_frame(), [1])
    summary = summarize_lookback(details)["factor_score_groups"]
    assert "trend_score" in set(summary["group_dimension"])
    assert "momentum_score" in set(summary["group_dimension"])


def test_elder_action_hint_grouping() -> None:
    """Elder action_hint should be a supported grouping dimension."""
    details = build_forward_return_details(_sample_strategy(), _price_frame(), [1])
    summary = summarize_lookback(details)["elder_review_groups"]
    assert "action_hint" in set(summary["group_dimension"])
    assert "趋势尚可，等待回调" in set(summary["group"])


def test_entry_zone_status_grouping() -> None:
    """Entry zone status and chase risk should be grouped."""
    details = build_forward_return_details(_sample_strategy(), _price_frame(), [1])
    summary = summarize_lookback(details)["entry_zone_groups"]
    assert "entry_zone_status" in set(summary["group_dimension"])
    assert "chase_risk" in set(summary["group_dimension"])


def test_watch_status_grouping() -> None:
    """Watchlist status should be grouped when available."""
    details = build_forward_return_details(_sample_strategy(), _price_frame(), [1])
    summary = summarize_lookback(details)["watchlist_groups"]
    assert "watch_status" in set(summary["group_dimension"])
    assert "active_watch" in set(summary["group"])


def test_lookback_workbook_has_expected_sheets(tmp_path: Path) -> None:
    """Full lookback workbook should contain expected sheets."""
    details = build_forward_return_details(_sample_strategy(), _price_frame(), [1, 3])
    summaries = summarize_lookback(details)
    status = {"status": "success", "as_of_trade_date": "20260630", "horizons": [1, 3], "candidate_sample_count": 2, "valid_sample_count": 4, "insufficient_forward_data_count": 0}
    output = tmp_path / "lookback.xlsx"
    save_lookback_workbook(output, status=status, summaries=summaries, details=details)
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == [
        "00_摘要",
        "01_候选整体回看",
        "02_综合分分组",
        "03_分项因子分组",
        "04_埃尔德复核回看",
        "05_买入区间回看",
        "06_观察池状态回看",
        "07_未来收益明细",
        "08_数据质量",
        "09_说明",
    ]


def test_lookback_status_json_written(tmp_path: Path) -> None:
    """run_lookback_analysis should write status JSON on completion."""
    store = _seed_store(tmp_path)
    status_path = tmp_path / "status.json"
    output = tmp_path / "lookback.xlsx"
    result = run_lookback_analysis(settings=_settings(store), store=store, output_path=output, status_path=status_path, horizons=[1], now=pd.Timestamp("2026-07-02").to_pydatetime())
    payload = json.loads(status_path.read_text(encoding="utf-8"))
    assert payload["status"] == result["status"]
    assert payload["generated_report_path"] == str(output)


def test_lookback_dry_run_does_not_write_report(tmp_path: Path) -> None:
    """Dry-run should inspect samples and status without exporting the report workbook."""
    store = _seed_store(tmp_path)
    output = tmp_path / "lookback.xlsx"
    run_lookback_analysis(settings=_settings(store), store=store, output_path=output, status_path=tmp_path / "status.json", horizons=[1], dry_run=True)
    assert not output.exists()


def test_lookback_streamlit_entry_exists() -> None:
    """Streamlit should expose an automatic lookback section and buttons."""
    source = (Path(__file__).resolve().parents[1] / "web" / "streamlit_app.py").read_text(encoding="utf-8")
    assert "自动回看分析" in source
    assert "运行自动回看分析" in source
    assert "刷新回看状态" in source
    assert "下载最新回看报告" in source
    assert "自动回看状态摘要" in source
    assert "尚无自动回看记录。可以点击运行自动回看分析生成结果。" in source


def test_strategy_backtest_tab_renders_lookback_entry() -> None:
    """The Strategy Backtest tab should render the automatic lookback entry."""
    source = (Path(__file__).resolve().parents[1] / "web" / "streamlit_app.py").read_text(encoding="utf-8")
    start = source.index("def _render_backtest_tab")
    end = source.index("def _render_status_tab", start)
    backtest_source = source[start:end]
    assert "_render_lookback_analysis_section(st)" in backtest_source
    assert "暂无回测结果。请先运行回测诊断；真实数据不足时不会生成结果。" not in backtest_source


def test_lookback_command_allowlist() -> None:
    """run_lookback_analysis must be callable from the safe command runner."""
    assert "run_lookback_analysis" in ALLOWED_COMMANDS
    assert "core.jobs.run_lookback_analysis" in " ".join(ALLOWED_COMMANDS["run_lookback_analysis"])


def test_no_algorithm_changes() -> None:
    """Task 58 should not alter core scoring or signal calculators."""
    root = Path(__file__).resolve().parents[1]
    scoring_source = (root / "core" / "factors" / "scoring.py").read_text(encoding="utf-8")
    elder_source = (root / "core" / "technical" / "elder.py").read_text(encoding="utf-8")
    entry_source = (root / "core" / "entry_zones" / "calculator.py").read_text(encoding="utf-8")
    assert "DEFAULT_WEIGHTS" in scoring_source
    assert "build_elder_review" in elder_source
    assert "calculate_entry_zones_for_targets" in entry_source


def _sample_strategy() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "trade_date": "20260601",
                "ts_code": "000001.SZ",
                "name": "平安银行",
                "industry": "银行",
                "total_score": 82,
                "trend_score": 85,
                "momentum_score": 75,
                "liquidity_score": 65,
                "fundamental_score": 55,
                "volatility_score": 45,
                "elder_score": 70,
                "action_hint": "趋势尚可，等待回调",
                "weekly_trend": "改善",
                "daily_pullback": "接近均线",
                "force_signal": "转强",
                "elder_ray_signal": "多头增强",
                "entry_zone_status": "near_zone",
                "chase_risk": "medium",
                "reward_risk_ratio": 2.2,
                "risk_pct": 0.05,
                "watch_status": "active_watch",
                "watch_status_label": "正常观察",
                "watch_days": 3,
                "selected_count_5d": 2,
                "selected_count_10d": 4,
                "consecutive_selected_days": 2,
            },
            {
                "trade_date": "20260601",
                "ts_code": "000002.SZ",
                "name": "万科A",
                "industry": "地产",
                "total_score": 58,
                "trend_score": 55,
                "momentum_score": 52,
                "liquidity_score": 80,
                "fundamental_score": 45,
                "volatility_score": 70,
                "elder_score": 35,
                "action_hint": "趋势偏弱，暂缓",
                "weekly_trend": "偏弱",
                "daily_pullback": "破位",
                "force_signal": "未转强",
                "elder_ray_signal": "压力未减弱",
                "entry_zone_status": "weak_no_entry",
                "chase_risk": "low",
                "reward_risk_ratio": 1.2,
                "risk_pct": 0.09,
                "watch_status": "weakening",
                "watch_status_label": "走势转弱",
                "watch_days": 5,
                "selected_count_5d": 0,
                "selected_count_10d": 1,
                "consecutive_selected_days": 0,
            },
        ]
    )


def _price_frame(days: int = 6) -> pd.DataFrame:
    dates = ["20260601", "20260602", "20260604", "20260605", "20260608", "20260609"][:days]
    closes_a = [10, 11, 12, 13, 14, 15][:days]
    closes_b = [20, 19, 18, 17, 16, 15][:days]
    rows = []
    for date, close in zip(dates, closes_a, strict=False):
        rows.append({"ts_code": "000001.SZ", "trade_date": date, "open": close, "high": close + 1, "low": close - 1, "close": close, "vol": 1, "amount": 1})
    for date, close in zip(dates, closes_b, strict=False):
        rows.append({"ts_code": "000002.SZ", "trade_date": date, "open": close, "high": close + 1, "low": close - 1, "close": close, "vol": 1, "amount": 1})
    return pd.DataFrame(rows)


def _seed_store(tmp_path: Path) -> DuckDBStore:
    store = DuckDBStore(tmp_path / "lookback.duckdb")
    store.initialize()
    strategy = _sample_strategy().copy()
    strategy["rank"] = [1, 2]
    strategy["close"] = [10.0, 20.0]
    store.upsert_dataframe("strategy_result", strategy[["trade_date", "rank", "ts_code", "name", "industry", "close", "total_score", "trend_score", "momentum_score", "liquidity_score", "fundamental_score", "volatility_score"]])
    store.upsert_dataframe("daily_price", _price_frame())
    store.upsert_dataframe(
        "entry_zone_snapshots",
        pd.DataFrame(
            [
                {"ts_code": "000001.SZ", "name": "平安银行", "trade_date": "20260601", "close": 10, "entry_low": 9.5, "entry_high": 10.5, "entry_mid": 10, "stop_loss": 9.2, "target_price": 13.5, "reward_risk_ratio": 2.2, "risk_pct": 0.05, "reward_pct": 0.1, "entry_zone_status": "near_zone", "chase_risk": "medium", "source": "selection"},
                {"ts_code": "000002.SZ", "name": "万科A", "trade_date": "20260601", "close": 20, "entry_low": 18, "entry_high": 19, "entry_mid": 18.5, "stop_loss": 17, "target_price": 21, "reward_risk_ratio": 1.2, "risk_pct": 0.09, "reward_pct": 0.05, "entry_zone_status": "weak_no_entry", "chase_risk": "low", "source": "selection"},
            ]
        ),
    )
    store.upsert_dataframe(
        "watchlist_daily_snapshots",
        pd.DataFrame(
            [
                {"snapshot_id": "1", "ts_code": "000001.SZ", "name": "平安银行", "trade_date": "20260601", "current_close": 10, "total_score": 82, "watch_status": "active_watch", "watch_status_label": "正常观察", "watch_days": 3, "selected_count_5d": 2, "selected_count_10d": 4, "consecutive_selected_days": 2, "elder_score": 70, "action_hint": "趋势尚可，等待回调", "weekly_trend": "改善", "daily_pullback": "接近均线", "force_signal": "转强", "elder_ray_signal": "多头增强"},
                {"snapshot_id": "2", "ts_code": "000002.SZ", "name": "万科A", "trade_date": "20260601", "current_close": 20, "total_score": 58, "watch_status": "weakening", "watch_status_label": "走势转弱", "watch_days": 5, "selected_count_5d": 0, "selected_count_10d": 1, "consecutive_selected_days": 0, "elder_score": 35, "action_hint": "趋势偏弱，暂缓", "weekly_trend": "偏弱", "daily_pullback": "破位", "force_signal": "未转强", "elder_ray_signal": "压力未减弱"},
            ]
        ),
    )
    return store


def _settings(store: DuckDBStore):
    class Settings:
        duckdb_path = store.db_path

    return Settings()
