"""Task 59B daily research Elder field simplification tests."""

from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook
import pandas as pd

from core.factors.scoring import DEFAULT_WEIGHTS
from core.jobs.export_daily_research_workbook import export_daily_research_workbook
from core.storage.duckdb_store import DuckDBStore
from tests.test_daily_research_workbook import _seed_store, _settings


REMOVED_SHEETS = {
    "02_埃尔德复核",
    "05_观察池跟踪",
    "07_风险提示",
    "08_数据质量",
    "09_参数配置",
    "10_说明",
}

ELDER_HEADERS = {
    "埃尔德分（elder_score）",
    "操作提示（action_hint）",
    "复核原因（elder_reason）",
    "周线趋势（weekly_trend）",
    "日线回调（daily_pullback）",
    "强力指数信号（force_signal）",
    "埃尔德射线信号（elder_ray_signal）",
}


def test_daily_research_removed_standalone_elder_and_diagnostic_sheets(tmp_path: Path) -> None:
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    workbook = load_workbook(output)
    assert not (REMOVED_SHEETS & set(workbook.sheetnames))


def test_candidate_sheet_contains_elder_fields_and_scores(tmp_path: Path) -> None:
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    sheet = load_workbook(output)["01_今日候选"]
    headers = _headers(sheet)
    assert ELDER_HEADERS <= set(headers)
    elder_idx = headers.index("埃尔德分（elder_score）") + 1
    assert all(sheet.cell(row=row, column=elder_idx).value is not None for row in range(2, sheet.max_row + 1))


def test_current_watchlist_sheet_contains_elder_fields_and_scores(tmp_path: Path) -> None:
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    sheet = load_workbook(output)["04_观察池"]
    headers = _headers(sheet)
    assert ELDER_HEADERS <= set(headers)
    elder_idx = headers.index("埃尔德分（elder_score）") + 1
    assert all(sheet.cell(row=row, column=elder_idx).value is not None for row in range(2, sheet.max_row + 1))


def test_missing_elder_fields_do_not_break_daily_research(tmp_path: Path) -> None:
    store = _seed_store_without_elder(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    workbook = load_workbook(output)
    assert "01_今日候选" in workbook.sheetnames
    assert "04_观察池" in workbook.sheetnames


def test_empty_watchlist_does_not_break_daily_research(tmp_path: Path) -> None:
    store = _seed_store(tmp_path)
    with store.connect() as connection:
        connection.execute("DELETE FROM watchlist_daily_snapshots")
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    sheet = load_workbook(output)["04_观察池"]
    assert "暂无观察池跟踪数据。" in str(sheet["A2"].value)


def test_current_watchlist_sheet_limits_to_30_rows(tmp_path: Path) -> None:
    store = _seed_store(tmp_path)
    rows = []
    for index in range(35):
        code = f"{index + 1:06d}.SZ"
        rows.append(
            {
                "snapshot_id": f"watch-{index}",
                "ts_code": code,
                "name": f"观察{index}",
                "trade_date": "20260630",
                "current_close": 10 + index,
                "total_score": 50 + index,
                "watch_status": "active_watch",
                "watch_status_label": "正常观察",
                "elder_score": 60 + index,
                "action_hint": "趋势尚可，等待回调",
                "elder_reason": "节奏复核",
                "weekly_trend": "改善",
                "daily_pullback": "接近均线",
                "force_signal": "转强",
                "elder_ray_signal": "多头改善",
            }
        )
    store.upsert_dataframe("watchlist_daily_snapshots", pd.DataFrame(rows))
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    sheet = load_workbook(output)["04_观察池"]
    assert sheet.max_row - 1 <= 30


def test_streamlit_no_standalone_elder_review_tab() -> None:
    source = Path("web/streamlit_app.py").read_text(encoding="utf-8")
    tabs_match = re.search(r"st\.tabs\(\[(.*?)\]\)", source, flags=re.S)
    assert tabs_match is not None
    assert "埃尔德复核" not in tabs_match.group(1)
    assert "Elder 复核已作为今日候选和当前观察池的附加判断字段展示" in source


def test_candidate_order_and_total_score_unchanged() -> None:
    assert DEFAULT_WEIGHTS == {
        "trend_score": 0.30,
        "momentum_score": 0.20,
        "liquidity_score": 0.20,
        "fundamental_score": 0.15,
        "volatility_score": 0.15,
    }
    selector_source = Path("core/strategy/selector.py").read_text(encoding="utf-8")
    assert 'sort_values(["trade_date", "total_score", "ts_code"], ascending=[True, False, True])' in selector_source


def _seed_store_without_elder(tmp_path: Path) -> DuckDBStore:
    store = _seed_store(tmp_path)
    strategy = store.read_table("strategy_result")
    elder_columns = [column for column in strategy.columns if column in {"elder_score", "action_hint", "elder_reason", "weekly_trend", "daily_pullback", "force_signal", "elder_ray_signal"}]
    if elder_columns:
        strategy = strategy.drop(columns=elder_columns)
        with store.connect() as connection:
            connection.execute("DELETE FROM strategy_result")
        store.upsert_dataframe("strategy_result", strategy)
    return store


def _headers(sheet) -> list[str]:
    return [cell.value for cell in sheet[1] if cell.value is not None]
