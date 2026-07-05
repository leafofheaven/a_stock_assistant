"""Task 59 watchlist lifecycle and Elder review scope tests."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
import pandas as pd

from core.jobs.export_daily_research_workbook import export_daily_research_workbook
from core.jobs.refresh_watchlist_from_selection import refresh_watchlist_from_selection
from core.jobs.run_elder_review import run_elder_review
from core.review.decisions import REVIEW_COLUMNS, update_review_decision
from core.review.tracking import read_watchlist_daily_snapshots
from core.storage.duckdb_store import DuckDBStore


def test_watchlist_does_not_append_duplicate_active_rows(tmp_path: Path) -> None:
    store = _seed_selection_store(tmp_path, symbol_count=3)
    settings = _settings(store, top_n=3)

    refresh_watchlist_from_selection(trade_date="20240105", top_n=3, quiet=True, settings=settings, store=store)
    refresh_watchlist_from_selection(trade_date="20240105", top_n=3, quiet=True, settings=settings, store=store)
    decisions = store.read_table("review_decisions")

    active = decisions[(decisions["decision"] == "watch") & (decisions["review_status"].isin(["active", "entry_zone", "triggered"]))]
    assert active["ts_code"].value_counts().max() == 1


def test_watchlist_daily_new_limit(tmp_path: Path) -> None:
    store = _seed_selection_store(tmp_path, symbol_count=30)
    settings = _settings(store, top_n=30)

    result = refresh_watchlist_from_selection(
        trade_date="20240105",
        top_n=30,
        daily_new_limit=20,
        quiet=True,
        settings=settings,
        store=store,
    )

    assert result["new_candidate_count"] == 20
    assert result["active_watch_count"] == 20


def test_watchlist_active_size_limit(tmp_path: Path) -> None:
    store = _seed_selection_store(tmp_path, symbol_count=5)
    _seed_existing_watch_decisions(store, count=65, status="active", selection_date="20240101")
    settings = _settings(store, top_n=5)

    result = refresh_watchlist_from_selection(
        trade_date="20240105",
        top_n=5,
        active_limit=60,
        quiet=True,
        settings=settings,
        store=store,
    )
    decisions = store.read_table("review_decisions")

    assert result["active_watch_count"] <= 60
    assert (decisions["review_status"] == "archived").sum() >= 5


def test_watchlist_expires_stale_symbols(tmp_path: Path) -> None:
    store = DuckDBStore(tmp_path / "stale.duckdb")
    store.initialize()
    store.upsert_dataframe(
        "strategy_result",
        pd.DataFrame(
            [
                {"trade_date": "20240101", "rank": 1, "ts_code": "000001.SZ", "name": "股票1", "total_score": 90},
                {"trade_date": "20240110", "rank": 1, "ts_code": "000002.SZ", "name": "股票2", "total_score": 95},
            ]
        ),
    )
    store.upsert_dataframe(
        "daily_price",
        pd.DataFrame(
            [
                {"trade_date": "20240101", "ts_code": "000001.SZ", "open": 10, "high": 11, "low": 9, "close": 10, "vol": 1000},
                {"trade_date": "20240110", "ts_code": "000002.SZ", "open": 10, "high": 11, "low": 9, "close": 10, "vol": 1000},
            ]
        ),
    )
    _seed_existing_watch_decisions(store, count=1, status="active", selection_date="20240101")
    settings = _settings(store, top_n=1)

    refresh_watchlist_from_selection(
        trade_date="20240110",
        top_n=1,
        stale_trading_days=5,
        quiet=True,
        settings=settings,
        store=store,
    )
    decisions = store.read_table("review_decisions")

    assert decisions.loc[decisions["ts_code"] == "000001.SZ", "review_status"].iloc[-1] == "expired"


def test_watchlist_keeps_entry_zone_priority(tmp_path: Path) -> None:
    store = _seed_selection_store(tmp_path, symbol_count=5)
    _seed_existing_watch_decisions(store, count=62, status="active", selection_date="20240101")
    _set_review_status(store, "000061.SZ", "entry_zone")
    _set_review_status(store, "000062.SZ", "triggered")
    settings = _settings(store, top_n=5)

    refresh_watchlist_from_selection(
        trade_date="20240105",
        top_n=5,
        active_limit=60,
        quiet=True,
        settings=settings,
        store=store,
    )
    decisions = store.read_table("review_decisions")
    statuses = dict(zip(decisions["ts_code"], decisions["review_status"], strict=False))

    assert statuses["000061.SZ"] == "entry_zone"
    assert statuses["000062.SZ"] == "triggered"


def test_watchlist_summary_counts_new_refreshed_exited(tmp_path: Path) -> None:
    store = _seed_selection_store(tmp_path, symbol_count=5)
    _seed_existing_watch_decisions(store, count=2, status="active", selection_date="20240101")
    settings = _settings(store, top_n=5)

    result = refresh_watchlist_from_selection(trade_date="20240105", top_n=5, quiet=True, settings=settings, store=store)

    assert result["new_candidate_count"] == 3
    assert result["refreshed_watch_count"] >= 2
    assert "exited_watch_count" in result


def test_elder_review_includes_active_watchlist(tmp_path: Path) -> None:
    store = _seed_elder_store(tmp_path)
    settings = _settings(store, top_n=1)
    update_review_decision(store=store, ts_code="000002.SZ", decision="watch", selection_date="20240101")

    result = run_elder_review(settings=settings, store=store, top_n=1, use_sample=False)
    review = result["elder_review_df"]

    assert "000002.SZ" in set(review["ts_code"])
    assert result["watchlist_review_count"] == 1
    assert "观察池" in set(review["review_scope"])


def test_elder_review_excludes_archived_watchlist_with_reason(tmp_path: Path) -> None:
    store = _seed_elder_store(tmp_path)
    settings = _settings(store, top_n=1)
    update_review_decision(store=store, ts_code="000002.SZ", decision="watch", selection_date="20240101")
    _set_review_status(store, "000002.SZ", "archived")

    result = run_elder_review(settings=settings, store=store, top_n=1, use_sample=False)
    review = result["elder_review_df"]

    assert "000002.SZ" not in set(review["ts_code"])
    assert "当前观察池" in "\n".join(result["notes"])


def test_elder_review_no_blank_score_without_reason(tmp_path: Path) -> None:
    store = _seed_elder_store(tmp_path, days=10)
    settings = _settings(store, top_n=1)

    review = run_elder_review(settings=settings, store=store, top_n=1, use_sample=False)["elder_review_df"]

    missing_score = pd.to_numeric(review["elder_score"], errors="coerce").isna() | (review["elder_score"] == 0)
    assert review.loc[missing_score, "review_reason"].fillna("").astype(str).str.len().min() > 0


def test_elder_review_does_not_change_candidate_order(tmp_path: Path) -> None:
    store = _seed_elder_store(tmp_path)
    settings = _settings(store, top_n=1)
    update_review_decision(store=store, ts_code="000002.SZ", decision="watch", selection_date="20240101")

    review = run_elder_review(settings=settings, store=store, top_n=1, use_sample=False)["elder_review_df"]

    assert review.iloc[0]["ts_code"] == "000001.SZ"
    assert review.iloc[0]["review_scope"] == "今日候选"


def test_daily_research_excel_watchlist_current_scope(tmp_path: Path) -> None:
    store = _seed_workbook_watchlist_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)
    sheet = load_workbook(output)["04_观察池"]
    values = _sheet_values(sheet)

    assert "000001.SZ" in values
    assert "000002.SZ" not in values


def test_daily_research_excel_elder_review_scope(tmp_path: Path) -> None:
    store = _seed_workbook_watchlist_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)
    sheet = load_workbook(output)["02_埃尔德复核"]
    values = _sheet_values(sheet)

    assert "今日候选" in values
    assert "观察池" in values
    assert "暂无埃尔德复核分；该行未找到可用复核结果或数据样本不足。" in values


def _settings(store: DuckDBStore, top_n: int = 10) -> SimpleNamespace:
    return SimpleNamespace(data_provider="akshare", duckdb_path=store.db_path, default_top_n=top_n)


def _seed_selection_store(tmp_path: Path, symbol_count: int, trade_dates: list[str] | None = None) -> DuckDBStore:
    store = DuckDBStore(tmp_path / "task59.duckdb")
    store.initialize()
    dates = trade_dates or ["20240105"]
    rows = []
    price_rows = []
    for trade_date in dates:
        for index in range(symbol_count):
            ts_code = f"{index + 1:06d}.SZ"
            rank = index + 1
            rows.append({"trade_date": trade_date, "rank": rank, "ts_code": ts_code, "name": f"股票{rank}", "total_score": 100 - rank})
            price_rows.append({"trade_date": trade_date, "ts_code": ts_code, "open": 10, "high": 11, "low": 9, "close": 10 + index, "vol": 1000, "amount": 100000})
    store.upsert_dataframe("strategy_result", pd.DataFrame(rows))
    store.upsert_dataframe("daily_price", pd.DataFrame(price_rows))
    return store


def _seed_existing_watch_decisions(store: DuckDBStore, count: int, status: str, selection_date: str) -> None:
    now = "2024-01-01T00:00:00"
    rows = []
    for index in range(count):
        ts_code = f"{index + 1:06d}.SZ"
        rows.append(
            {
                "decision_id": f"watch-{ts_code}",
                "ts_code": ts_code,
                "name": f"股票{index + 1}",
                "selection_date": selection_date,
                "review_date": selection_date,
                "decision": "watch",
                "review_status": status,
                "reviewer": "test",
                "reason": "seed",
                "notes": "",
                "data_quality_note": "",
                "source_report_path": "",
                "created_at": now,
                "updated_at": now,
            }
        )
    store.upsert_dataframe("review_decisions", pd.DataFrame(rows, columns=REVIEW_COLUMNS))


def _set_review_status(store: DuckDBStore, ts_code: str, status: str) -> None:
    decisions = store.read_table("review_decisions")
    decisions.loc[decisions["ts_code"] == ts_code, "review_status"] = status
    store.upsert_dataframe("review_decisions", decisions)


def _seed_elder_store(tmp_path: Path, days: int = 70) -> DuckDBStore:
    store = DuckDBStore(tmp_path / "elder_task59.duckdb")
    store.initialize()
    store.upsert_dataframe(
        "stock_basic",
        pd.DataFrame(
            [
                {"ts_code": "000001.SZ", "symbol": "000001", "name": "平安银行", "industry": "银行", "market": "主板"},
                {"ts_code": "000002.SZ", "symbol": "000002", "name": "万科A", "industry": "地产", "market": "主板"},
            ]
        ),
    )
    store.upsert_dataframe(
        "strategy_result",
        pd.DataFrame(
            [
                {"trade_date": "20240430", "rank": 1, "ts_code": "000001.SZ", "name": "平安银行", "industry": "银行", "total_score": 90.0},
                {"trade_date": "20240430", "rank": 2, "ts_code": "000002.SZ", "name": "万科A", "industry": "地产", "total_score": 80.0},
            ]
        ),
    )
    store.upsert_dataframe("daily_price", _price_frame(days=days))
    return store


def _price_frame(days: int) -> pd.DataFrame:
    rows = []
    dates = pd.bdate_range("2024-01-01", periods=days).strftime("%Y%m%d")
    for ts_code, base in [("000001.SZ", 10.0), ("000002.SZ", 12.0)]:
        for index, trade_date in enumerate(dates):
            close = base + index * 0.05
            rows.append({"ts_code": ts_code, "trade_date": trade_date, "open": close, "high": close * 1.02, "low": close * 0.98, "close": close, "vol": 1_000_000})
    return pd.DataFrame(rows)


def _seed_workbook_watchlist_store(tmp_path: Path) -> DuckDBStore:
    store = _seed_elder_store(tmp_path)
    store.upsert_dataframe(
        "watchlist_daily_snapshots",
        pd.DataFrame(
            [
                {"snapshot_id": "1", "trade_date": "20240430", "ts_code": "000001.SZ", "name": "平安银行", "watch_status": "active", "watch_status_label": "正常观察", "elder_score": pd.NA, "elder_reason": ""},
                {"snapshot_id": "2", "trade_date": "20240430", "ts_code": "000002.SZ", "name": "万科A", "watch_status": "expired", "watch_status_label": "观察过期", "elder_score": 55, "elder_reason": "历史记录"},
            ]
        ),
    )
    return store


def _sheet_values(sheet) -> list[object]:
    return [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
