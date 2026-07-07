"""Task 59C daily research buy-zone and lookback status tests."""

from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
import pandas as pd

from core.jobs.export_daily_research_workbook import export_daily_research_workbook
from core.jobs.export_daily_research_workbook import build_lookback_status_display
from core.storage.duckdb_store import DuckDBStore
from core.review.decisions import REVIEW_COLUMNS


def test_export_autocalculates_same_day_entry_zones_when_missing(tmp_path: Path) -> None:
    store = _seed_research_scope_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=tmp_path / "missing.json")

    workbook = load_workbook(output)
    rows = _sheet_records(workbook["03_买入区间"])
    assert rows
    assert {row["交易日期（trade_date）"] for row in rows} == {"20260706"}


def test_entry_zone_sheet_uses_visible_candidate_and_watchlist_scope(tmp_path: Path) -> None:
    store = _seed_research_scope_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=tmp_path / "missing.json")

    workbook = load_workbook(output)
    candidate_codes = {row["股票代码（ts_code）"] for row in _sheet_records(workbook["01_今日候选"])}
    watchlist_codes = {row["股票代码（ts_code）"] for row in _sheet_records(workbook["04_观察池"])}
    entry_rows = _sheet_records(workbook["03_买入区间"])
    entry_codes = {row["股票代码（ts_code）"] for row in entry_rows}
    assert len(_sheet_records(workbook["04_观察池"])) == 30
    assert entry_codes == candidate_codes | watchlist_codes
    assert len(entry_rows) == len(entry_codes)
    assert len(entry_rows) <= 40
    assert len(entry_rows) != 62
    for row in entry_rows:
        expected_source = "selection" if row["股票代码（ts_code）"] in candidate_codes else "watchlist"
        assert row["来源（source）"] == expected_source


def test_entry_zone_duplicate_prefers_selection_source(tmp_path: Path) -> None:
    store = _seed_research_scope_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=tmp_path / "missing.json")

    rows = _sheet_records(load_workbook(output)["03_买入区间"])
    duplicate = next(row for row in rows if row["股票代码（ts_code）"] == "000001.SZ")
    assert duplicate["来源（source）"] == "selection"


def test_stale_lookback_creates_refresh_status_sheet(tmp_path: Path) -> None:
    store = _seed_research_scope_store(tmp_path)
    status_path = _write_lookback_status(tmp_path, "20260703", valid_sample_count=20)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    workbook = load_workbook(output)
    assert "11_自动回看摘要" not in workbook.sheetnames
    assert "11_自动回看状态" in workbook.sheetnames
    values = "\n".join(str(value) for value in _sheet_values(workbook["11_自动回看状态"]))
    assert "需要刷新当日回看" in values
    assert ".venv/bin/python -m core.jobs.run_lookback_analysis --as-of 20260706 --format text" in values


def test_same_day_valid_lookback_keeps_summary_sheet(tmp_path: Path) -> None:
    store = _seed_research_scope_store(tmp_path)
    status_path = _write_lookback_status(tmp_path, "20260706", valid_sample_count=20)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    workbook = load_workbook(output)
    assert "11_自动回看摘要" in workbook.sheetnames
    assert "11_自动回看状态" not in workbook.sheetnames


def test_status_page_uses_latest_local_research_date_when_planned_date_has_zero_coverage(monkeypatch) -> None:
    import web.streamlit_app as app

    captured: dict[str, str] = {}

    def fake_snapshot(*, db_path: str, research_trade_date: str, latest_completed_trade_date: str) -> dict[str, object]:
        captured["research_trade_date"] = research_trade_date
        return {
            "data_quality_status": "ok",
            "configured_symbol_count": 5055,
            "latest_daily_price_symbol_count": 4994,
            "latest_completed_trade_date": latest_completed_trade_date,
        }

    monkeypatch.setattr(app, "build_data_quality_snapshot", fake_snapshot)
    scheduled = {
        "data_quality_snapshot_source": "scheduled",
        "latest_completed_trade_date": "20260707",
        "latest_daily_price_symbol_count": 0,
    }
    legacy = {"latest_trade_date": "20260706", "latest_selection_date": "20260706", "duckdb_path": "/tmp/test.duckdb"}

    status = app._status_page_quality_snapshot({"_duckdb_path": "/tmp/test.duckdb"}, scheduled, legacy)

    assert captured["research_trade_date"] == "20260706"
    assert status["current_research_trade_date"] == "20260706"
    assert status["planned_update_target_date"] == "20260707"
    assert status["latest_daily_price_symbol_count"] == 4994
    assert "计划目标日期 20260707 尚未完成更新" in status["formal_result_warning_reason"]


def test_status_page_uses_current_research_date_even_when_planned_snapshot_has_rows(monkeypatch) -> None:
    import web.streamlit_app as app

    captured: dict[str, str] = {}

    def fake_snapshot(*, db_path: str, research_trade_date: str, latest_completed_trade_date: str) -> dict[str, object]:
        captured["research_trade_date"] = research_trade_date
        return {
            "data_quality_status": "ok",
            "configured_symbol_count": 5055,
            "latest_daily_price_symbol_count": 4994,
            "latest_completed_trade_date": latest_completed_trade_date,
        }

    monkeypatch.setattr(app, "build_data_quality_snapshot", fake_snapshot)
    scheduled = {
        "data_quality_snapshot_source": "scheduled",
        "latest_completed_trade_date": "20260707",
        "latest_daily_price_symbol_count": 68,
    }
    legacy = {"latest_trade_date": "20260706", "latest_selection_date": "20260706", "duckdb_path": "/tmp/test.duckdb"}

    status = app._status_page_quality_snapshot({"_duckdb_path": "/tmp/test.duckdb"}, scheduled, legacy)

    assert captured["research_trade_date"] == "20260706"
    assert status["latest_daily_price_symbol_count"] == 4994
    assert status["current_research_trade_date"] == "20260706"
    assert status["planned_update_target_date"] == "20260707"


def test_streamlit_daily_research_frames_match_excel_scope(tmp_path: Path) -> None:
    import web.streamlit_app as app

    store = _seed_research_scope_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"
    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=tmp_path / "missing.json")

    tables = {
        "strategy_result": store.read_table("strategy_result"),
        "entry_zone_snapshots": store.read_table("entry_zone_snapshots"),
        "_watchlist_snapshot": store.read_table("watchlist_daily_snapshots"),
        "external_position_snapshots": pd.DataFrame(),
        "daily_price": store.read_table("daily_price"),
    }
    view = app._build_dashboard_daily_research_view(tables, tables["daily_price"])
    workbook = load_workbook(output)

    excel_candidates = {row["股票代码（ts_code）"] for row in _sheet_records(workbook["01_今日候选"])}
    excel_watchlist = {row["股票代码（ts_code）"] for row in _sheet_records(workbook["04_观察池"])}
    excel_entry = {row["股票代码（ts_code）"] for row in _sheet_records(workbook["03_买入区间"])}

    assert view is not None
    assert set(view.strategy_sheet["ts_code"]) == excel_candidates
    assert set(view.watchlist_sheet["ts_code"]) == excel_watchlist
    assert set(view.entry_sheet["ts_code"]) == excel_entry
    assert view.strategy_sheet["ts_code"].is_unique
    assert len(view.watchlist_sheet) == 30
    assert len(view.entry_sheet) == len(excel_entry)
    assert len(view.entry_sheet) <= 40


def test_streamlit_lookback_status_uses_shared_display_function() -> None:
    import web.streamlit_app as app

    status = {
        "as_of_trade_date": "20260703",
        "end_date": "20260703",
        "candidate_sample_count": 30,
        "valid_sample_count": 18,
        "insufficient_forward_data_count": 232,
        "generated_report_path": "/tmp/missing_lookback.xlsx",
    }

    assert app.build_lookback_status_display is build_lookback_status_display
    display = app.build_lookback_status_display(status, "20260706")
    assert display["is_current"] is False
    assert display["summary"]["回看状态"] == "需要刷新当日回看"
    assert display["summary"]["报告文件状态"] == "文件不存在，可能已清理"


def _seed_research_scope_store(tmp_path: Path) -> DuckDBStore:
    store = DuckDBStore(tmp_path / "task59c.duckdb")
    store.initialize()
    codes = [f"{index:06d}.SZ" for index in range(1, 63)]
    store.upsert_dataframe(
        "stock_basic",
        pd.DataFrame(
            [
                {"ts_code": code, "symbol": code.split(".")[0], "name": f"股票{idx}", "industry": "测试", "market": "主板", "exchange": "SZSE"}
                for idx, code in enumerate(codes, start=1)
            ]
        ),
    )
    dates = pd.bdate_range(end="2026-07-06", periods=40).strftime("%Y%m%d").tolist()
    price_rows = []
    for idx, code in enumerate(codes, start=1):
        for day_index, trade_date in enumerate(dates):
            close = 10 + idx * 0.01 + day_index * 0.02
            price_rows.append(
                {
                    "ts_code": code,
                    "trade_date": trade_date,
                    "open": close,
                    "high": close * 1.02,
                    "low": close * 0.98,
                    "close": close,
                    "pre_close": close * 0.99,
                    "change": close * 0.01,
                    "pct_chg": 1.0,
                    "vol": 1_000_000,
                    "amount": 100_000_000,
                }
            )
    store.upsert_dataframe("daily_price", pd.DataFrame(price_rows))
    strategy_rows = [
        {
            "trade_date": "20260706",
            "rank": idx,
            "ts_code": code,
            "name": f"股票{idx}",
            "industry": "测试",
            "close": 10 + idx,
            "total_score": 100 - idx,
            "trend_score": 80,
            "momentum_score": 80,
            "liquidity_score": 80,
            "fundamental_score": 60,
            "volatility_score": 60,
        }
        for idx, code in enumerate(codes[:10], start=1)
    ]
    store.upsert_dataframe("strategy_result", pd.DataFrame(strategy_rows))
    watch_codes = [*codes[:5], *codes[10:57]]
    snapshot_rows = [
        {
            "snapshot_id": f"watch-{code}",
            "trade_date": "20260706",
            "ts_code": code,
            "name": f"观察{idx}",
            "current_close": 10 + idx,
            "total_score": 70 - idx * 0.1,
            "watch_status": "active_watch",
            "watch_status_label": "正常观察",
            "elder_score": 50 + idx,
            "action_hint": "趋势尚可，等待回调",
            "elder_reason": "测试",
        }
        for idx, code in enumerate(watch_codes, start=1)
    ]
    store.upsert_dataframe("watchlist_daily_snapshots", pd.DataFrame(snapshot_rows))
    now = "2026-07-06T18:00:00"
    decision_rows = [
        {
            "decision_id": f"decision-{code}",
            "ts_code": code,
            "name": f"观察{idx}",
            "selection_date": "20260706",
            "review_date": "20260706",
            "decision": "watch",
            "review_status": "active",
            "reviewer": "test",
            "reason": "测试",
            "notes": "",
            "data_quality_note": "",
            "source_report_path": "",
            "created_at": now,
            "updated_at": now,
        }
        for idx, code in enumerate(watch_codes, start=1)
    ]
    store.upsert_dataframe("review_decisions", pd.DataFrame(decision_rows, columns=REVIEW_COLUMNS))
    return store


def _settings(store: DuckDBStore) -> SimpleNamespace:
    return SimpleNamespace(duckdb_path=store.db_path, data_provider="akshare", real_universe_preset="full")


def _write_lookback_status(tmp_path: Path, as_of_trade_date: str, *, valid_sample_count: int) -> Path:
    path = tmp_path / f"lookback_{as_of_trade_date}.json"
    path.write_text(
        json.dumps(
            {
                "status": "success",
                "as_of_trade_date": as_of_trade_date,
                "start_date": "20260601",
                "end_date": as_of_trade_date,
                "horizons": [5, 10, 20],
                "candidate_sample_count": 30,
                "valid_sample_count": valid_sample_count,
                "insufficient_forward_data_count": 3,
                "total_score_group_summary": "有效摘要",
                "elder_review_summary": "有效摘要",
                "entry_zone_summary": "有效摘要",
                "watchlist_summary": "有效摘要",
                "key_findings": "有效发现",
                "generated_report_path": "/tmp/lookback.xlsx",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


def _sheet_records(sheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in sheet[1]]
    return [
        {headers[index]: value for index, value in enumerate(row)}
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in row)
    ]


def _sheet_values(sheet) -> list[object]:
    return [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
