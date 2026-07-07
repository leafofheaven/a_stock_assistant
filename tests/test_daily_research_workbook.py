"""Tests for daily research workbook export."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
import json

from openpyxl import load_workbook
import pandas as pd

from core.factors.scoring import DEFAULT_WEIGHTS
from core.jobs.export_daily_research_workbook import (
    SHEET_NAMES,
    WorkbookExportResult,
    _build_data_quality_sheet,
    _help_sheet,
    export_daily_research_workbook,
    main as export_workbook_main,
    _resolve_output_path,
    _settings_sheet,
)
from core.storage.duckdb_store import DuckDBStore


def test_export_daily_research_workbook_writes_required_sheets(tmp_path: Path) -> None:
    """Workbook export should include all required sheets and key rows."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    result = export_daily_research_workbook(
        output_path=output,
        settings=_settings(store),
        store=store,
        lookback_status_path=tmp_path / "missing_lookback.json",
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == SHEET_NAMES
    assert result.strategy_rows == 2
    assert result.entry_zone_rows == 2
    assert result.watchlist_rows == 1
    assert result.external_position_rows == 1


FORBIDDEN_RANK_HEADERS = {
    "rank",
    "candidate_rank",
    "original_rank",
    "today_rank",
    "previous_rank",
    "rank_change",
    "best_rank",
    "latest_rank",
    "原始选股排名",
    "当日入选排名",
    "上次入选排名",
    "排名变化",
    "历史最佳入选排名",
    "最近入选排名",
}


def test_workbook_has_no_rank_columns_by_default(tmp_path: Path) -> None:
    """Default workbook sheets should hide all rank-like fields."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    workbook = load_workbook(output)
    for sheet_name in workbook.sheetnames:
        headers = _headers(workbook[sheet_name])
        joined = "\n".join(str(header) for header in headers)
        for forbidden in FORBIDDEN_RANK_HEADERS:
            assert forbidden not in headers
            assert forbidden not in joined


def test_candidate_sheet_fields(tmp_path: Path) -> None:
    """Candidate sheet should use continuous display order and Chinese-first labels."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    sheet = load_workbook(output)["01_今日候选"]
    headers = _headers(sheet)
    display_idx = headers.index("序号") + 1
    code_idx = headers.index("股票代码（ts_code）") + 1

    assert [sheet.cell(row=row, column=display_idx).value for row in (2, 3)] == [1, 2]
    assert [sheet.cell(row=row, column=code_idx).value for row in (2, 3)] == ["000001.SZ", "000002.SZ"]
    for header in [
        "交易日期（trade_date）",
        "综合分（total_score）",
        "趋势分（trend_score）",
        "动量分（momentum_score）",
        "流动性分（liquidity_score）",
        "基本面分（fundamental_score）",
        "波动分（volatility_score）",
    ]:
        assert header in headers


def test_workbook_uses_chinese_name_with_english_field_labels(tmp_path: Path) -> None:
    """Important workbook fields should follow 中文名称（英文名） labels."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    workbook = load_workbook(output)
    all_headers = {header for sheet in workbook.worksheets for header in _headers(sheet)}
    for header in [
        "综合分（total_score）",
        "趋势分（trend_score）",
        "动量分（momentum_score）",
        "流动性分（liquidity_score）",
        "基本面分（fundamental_score）",
        "波动分（volatility_score）",
        "埃尔德分（elder_score）",
        "操作提示（action_hint）",
        "买入区间下限（entry_low）",
        "止损价（stop_loss）",
        "盈亏比（reward_risk_ratio）",
    ]:
        assert header in all_headers


def test_daily_research_workbook_includes_lookback_summary_when_status_exists(tmp_path: Path) -> None:
    """Daily workbook should include a lightweight lookback summary sheet when status exists."""
    store = _seed_store(tmp_path)
    status_path = _write_lookback_status(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    workbook = load_workbook(output)
    assert "11_自动回看摘要" in workbook.sheetnames
    sheet = workbook["11_自动回看摘要"]
    headers = _headers(sheet)
    assert "回看截止交易日（as_of_trade_date）" in headers
    assert "完整回看报告路径（lookback_report_path）" in headers
    values = [cell.value for cell in sheet[2]]
    assert "20260630" in values
    assert "/tmp/lookback.xlsx" in values


def test_daily_research_workbook_handles_missing_lookback_status(tmp_path: Path) -> None:
    """Missing lookback status should not break workbook export."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store, lookback_status_path=tmp_path / "missing.json")

    workbook = load_workbook(output)
    assert "11_自动回看摘要" not in workbook.sheetnames
    summary_values = _sheet_values(workbook["00_摘要"])
    assert "尚无自动回看记录。" in summary_values


def test_daily_research_workbook_does_not_embed_full_lookback_detail(tmp_path: Path) -> None:
    """Daily workbook should not embed the full future-return detail table."""
    store = _seed_store(tmp_path)
    status_path = _write_lookback_status(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    workbook = load_workbook(output)
    assert "07_未来收益明细" not in workbook.sheetnames
    assert "11_自动回看摘要" in workbook.sheetnames


def test_lookback_report_path_is_recorded_in_daily_workbook_summary(tmp_path: Path) -> None:
    """00 summary should record the latest full lookback report path."""
    store = _seed_store(tmp_path)
    status_path = _write_lookback_status(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    sheet = load_workbook(output)["00_摘要"]
    rows = {(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value) for i in range(2, sheet.max_row + 1)}
    assert ("完整回看报告路径", "/tmp/lookback.xlsx") in rows


def test_entry_zone_sheet_does_not_fallback_to_stale_trade_date(tmp_path: Path) -> None:
    """Entry zone rows must align to the workbook research date."""
    store = _seed_store(tmp_path)
    store.upsert_dataframe(
        "entry_zone_snapshots",
        pd.DataFrame(
            [
                {
                    "ts_code": "000001.SZ",
                    "name": "平安银行",
                    "trade_date": "20260703",
                    "close": 10.5,
                    "entry_low": 9.8,
                    "entry_high": 10.2,
                    "entry_mid": 10.0,
                    "stop_loss": 9.2,
                    "target_price": 11.6,
                    "reward_risk_ratio": 2.0,
                    "entry_zone_status": "near_zone",
                    "source": "selection",
                }
            ]
        ),
    )
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store)

    sheet = load_workbook(output)["03_买入区间"]
    assert "交易日期（trade_date）" in _headers(sheet)
    assert "20260703" not in _sheet_values(sheet)
    assert sheet.max_row == 1


def test_stale_lookback_status_is_not_expanded(tmp_path: Path) -> None:
    """Old lookback summaries should be skipped for a newer daily workbook."""
    store = _seed_store(tmp_path)
    status_path = _write_lookback_status(tmp_path, as_of_trade_date="20260703")
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(trade_date="20260706", output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    workbook = load_workbook(output)
    assert "11_自动回看摘要" not in workbook.sheetnames
    summary_values = _sheet_values(workbook["00_摘要"])
    assert "自动回看：最近回看截止 20260703，早于当前研究日期 20260706，本日报未展开旧回看摘要。" in summary_values


def test_uninformative_lookback_status_is_not_expanded(tmp_path: Path) -> None:
    """Lookback statuses with no usable samples should stay out of the daily workbook."""
    store = _seed_store(tmp_path)
    status_path = _write_lookback_status(
        tmp_path,
        as_of_trade_date="20260630",
        candidate_sample_count=0,
        valid_sample_count=0,
        total_score_group_summary="暂无可统计样本",
        elder_review_summary="暂无可统计样本",
        entry_zone_summary="暂无可统计样本",
        watchlist_summary="暂无可统计样本",
        key_findings="暂无可统计样本",
    )
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store, lookback_status_path=status_path)

    workbook = load_workbook(output)
    assert "11_自动回看摘要" not in workbook.sheetnames
    assert "自动回看：最近回看有效样本不足，本日报未展开回看摘要。" in _sheet_values(workbook["00_摘要"])


def test_export_log_uses_elder_field_wording(tmp_path: Path, capsys, monkeypatch) -> None:
    """CLI output should not imply a standalone Elder review sheet still exists."""
    output = tmp_path / "daily_research.xlsx"
    import core.jobs.export_daily_research_workbook as export_module

    def fake_export(**_: object) -> WorkbookExportResult:
        return WorkbookExportResult(
            output_path=output,
            trade_date="20260706",
            strategy_rows=10,
            elder_rows=40,
            entry_zone_rows=0,
            watchlist_rows=30,
            external_position_rows=0,
            lookback_summary_rows=0,
        )

    monkeypatch.setattr(export_module, "export_daily_research_workbook", fake_export)

    exit_code = export_workbook_main(["--output", str(output), "--lookback-status-path", str(tmp_path / "missing.json")])

    captured = capsys.readouterr().out
    assert exit_code == 0
    assert "埃尔德复核:" not in captured
    assert "Elder 字段补充:" in captured


def test_elder_review_sheet_removed(tmp_path: Path) -> None:
    """Elder review should be embedded into candidate/watchlist sheets, not a standalone sheet."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    workbook = load_workbook(output)
    assert "02_埃尔德复核" not in workbook.sheetnames
    candidate_headers = _headers(workbook["01_今日候选"])
    for header in ["埃尔德分（elder_score）", "操作提示（action_hint）", "复核原因（elder_reason）"]:
        assert header in candidate_headers


def test_watchlist_sheet_hides_rank_fields(tmp_path: Path) -> None:
    """Current watchlist sheet should hide rank fields."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    headers = _headers(load_workbook(output)["04_观察池"])
    assert "观察状态（watch_status）" in headers
    assert "近5日入选次数（selected_count_5d）" in headers
    assert not (FORBIDDEN_RANK_HEADERS & set(headers))


def test_watchlist_tracking_sheet_removed(tmp_path: Path) -> None:
    """Watchlist tracking details should not be exported as a main daily sheet."""
    store = _seed_store(tmp_path)
    output = tmp_path / "daily_research.xlsx"

    export_daily_research_workbook(output_path=output, settings=_settings(store), store=store)

    assert "05_观察池跟踪" not in load_workbook(output).sheetnames


def test_workbook_export_is_read_only_for_duckdb(tmp_path: Path) -> None:
    """Export should not write or recompute local strategy rows."""
    store = _seed_store(tmp_path)
    before = _table_counts(store)

    export_daily_research_workbook(
        output_path=tmp_path / "daily_research.xlsx",
        settings=_settings(store),
        store=store,
    )

    assert _table_counts(store) == before


def test_workbook_filters_sensitive_settings(tmp_path: Path) -> None:
    """Settings helper must not expose token/key/password/secret values."""
    store = _seed_store(tmp_path)
    settings = SimpleNamespace(
        duckdb_path=store.db_path,
        data_provider="akshare",
        real_universe_preset="full",
        tushare_token="SECRET_TOKEN",
        api_key="SECRET_KEY",
        password="SECRET_PASSWORD",
        extra_value="sk-proj-SECRET",
    )
    sheet = _settings_sheet(settings)
    joined = "\n".join(str(value) for value in sheet.to_numpy().flatten().tolist())
    assert "SECRET" not in joined
    assert "duckdb_path" in joined
    assert "real_universe_preset" in joined


def test_data_quality_missing_values_wording(tmp_path: Path) -> None:
    """Data-quality helper should not describe missing values as missing fields."""
    store = _seed_store(tmp_path)

    frame = _build_data_quality_sheet(store, "20260630")
    joined = "\n".join(str(value) for value in frame.to_numpy().flatten().tolist())
    assert "基本面分（fundamental_score）缺失记录数" in joined
    assert "字段缺失：fundamental_score" not in joined
    assert "字段缺失：{'fundamental_score'" not in joined


def test_explanation_sheet_defines_no_rank_policy(tmp_path: Path) -> None:
    """Explanation helper should explain sequence, field naming, and no-rank policy."""
    values = _help_sheet().to_numpy().flatten().tolist()
    joined = "\n".join(str(value) for value in values)
    assert "默认不导出 rank / 排名字段" in joined
    assert "序号只代表当前 Sheet 当前显示顺序，不代表买入优先级" in joined
    assert "中文名称（英文名）" in joined
    assert "04_观察池是当前观察名单" in joined
    assert "不是交易指令" in joined


def test_empty_database_still_exports_clear_workbook(tmp_path: Path) -> None:
    """Missing or empty local results should produce an Excel workbook with clear messages."""
    store = DuckDBStore(tmp_path / "empty.duckdb")
    output = tmp_path / "empty.xlsx"

    result = export_daily_research_workbook(
        output_path=output,
        settings=_settings(store),
        store=store,
    )

    sheet = load_workbook(output)["01_今日候选"]
    assert output.exists()
    assert result.strategy_rows == 1
    assert "暂无本地选股结果" in str(sheet["A2"].value)


def test_task53_verifier_uses_temp_output() -> None:
    """Task 53 verification should not write workbook output into reports/."""
    source = Path("scripts/verify_task.py").read_text(encoding="utf-8")

    assert "task53" in source
    assert "/tmp/a_stock_assistant_task53" in source
    assert "reports/daily_research" not in source


def test_default_workbook_filename_does_not_repeat_trade_date() -> None:
    """Default output should be daily_research_YYYYMMDD_HHMMSS.xlsx."""
    path = _resolve_output_path(None, "20260630")

    assert path.name.startswith("daily_research_20260630_")
    assert path.name.count("20260630") == 1
    assert path.suffix == ".xlsx"


def test_no_algorithm_changes() -> None:
    """Task 56 should not alter scoring, selection, Elder, or entry-zone logic."""
    assert DEFAULT_WEIGHTS == {
        "trend_score": 0.30,
        "momentum_score": 0.20,
        "liquidity_score": 0.20,
        "fundamental_score": 0.15,
        "volatility_score": 0.15,
    }
    root = Path(__file__).resolve().parents[1]
    selection_source = (root / "core" / "strategy" / "selector.py").read_text(encoding="utf-8")
    elder_source = (root / "core" / "technical" / "elder.py").read_text(encoding="utf-8")
    entry_zone_source = (root / "core" / "entry_zones" / "calculator.py").read_text(encoding="utf-8")
    assert 'sort_values(["trade_date", "total_score", "ts_code"], ascending=[True, False, True])' in selection_source
    assert "does not replace or\n    modify ``total_score``" in elder_source
    assert "calculate_entry_zones_for_targets" in entry_zone_source


def _headers(sheet) -> list[str]:
    return [cell.value for cell in sheet[1] if cell.value is not None]


def _sheet_values(sheet) -> list[object]:
    return [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]


def _seed_store(tmp_path: Path) -> DuckDBStore:
    store = DuckDBStore(tmp_path / "research.duckdb")
    store.initialize()
    store.upsert_dataframe(
        "stock_basic",
        pd.DataFrame(
            [
                {"ts_code": "000001.SZ", "symbol": "000001", "name": "平安银行", "industry": "银行", "market": "主板", "exchange": "SZSE"},
                {"ts_code": "000002.SZ", "symbol": "000002", "name": "万科A", "industry": "地产", "market": "主板", "exchange": "SZSE"},
            ]
        ),
    )
    store.upsert_dataframe(
        "daily_price",
        pd.DataFrame(
            [
                {"ts_code": "000001.SZ", "trade_date": "20260630", "open": 10, "high": 11, "low": 9, "close": 10.5, "pre_close": 10, "change": 0.5, "pct_chg": 5, "vol": 1, "amount": 1},
                {"ts_code": "000002.SZ", "trade_date": "20260630", "open": 20, "high": 21, "low": 19, "close": 20.5, "pre_close": 20, "change": 0.5, "pct_chg": 2.5, "vol": 1, "amount": 1},
            ]
        ),
    )
    store.upsert_dataframe(
        "factor_scores",
        pd.DataFrame(
            [
                {"ts_code": "000001.SZ", "trade_date": "20260630", "trend_score": 90, "momentum_score": 80, "liquidity_score": 70, "volatility_score": 60, "fundamental_score": 50, "total_score": 78},
                {"ts_code": "000002.SZ", "trade_date": "20260630", "trend_score": 70, "momentum_score": 60, "liquidity_score": 50, "volatility_score": 40, "fundamental_score": None, "total_score": 58},
            ]
        ),
    )
    store.upsert_dataframe(
        "strategy_result",
        pd.DataFrame(
            [
                {"trade_date": "20260630", "rank": 1, "ts_code": "000001.SZ", "name": "平安银行", "industry": "银行", "close": 10.5, "pe": 6.1, "pb": 0.8, "total_score": 78.0, "trend_score": 90.0, "momentum_score": 80.0, "liquidity_score": 70.0, "fundamental_score": 50.0, "volatility_score": 60.0},
                {"trade_date": "20260630", "rank": 2, "ts_code": "000002.SZ", "name": "万科A", "industry": "地产", "close": 20.5, "pe": 7.2, "pb": 0.9, "total_score": 58.0, "trend_score": 70.0, "momentum_score": 60.0, "liquidity_score": 50.0, "fundamental_score": 30.0, "volatility_score": 40.0},
            ]
        ),
    )
    store.upsert_dataframe(
        "entry_zone_snapshots",
        pd.DataFrame(
            [
                {"ts_code": "000001.SZ", "name": "平安银行", "trade_date": "20260630", "close": 10.5, "ema13": 10.0, "ema22": 9.8, "ema60": 9.0, "entry_low": 9.8, "entry_high": 10.2, "entry_mid": 10.0, "stop_loss": 9.2, "target_price": 11.6, "reward_risk_ratio": 2.0, "entry_zone_status": "near_zone", "entry_zone_status_cn": "接近买入区间", "chase_risk": "medium", "chase_risk_cn": "中", "source": "selection"},
                {"ts_code": "000002.SZ", "name": "万科A", "trade_date": "20260630", "close": 20.5, "ema13": 19.0, "ema22": 18.8, "ema60": 18.0, "entry_low": 18.8, "entry_high": 19.5, "entry_mid": 19.15, "stop_loss": 17.8, "target_price": 21.85, "reward_risk_ratio": 2.0, "entry_zone_status": "above_zone", "entry_zone_status_cn": "高于买入区间", "chase_risk": "high", "chase_risk_cn": "高", "source": "selection"},
            ]
        ),
    )
    store.upsert_dataframe(
        "watchlist_daily_snapshots",
        pd.DataFrame(
            [
                {
                    "snapshot_id": "watch-000001-20260630",
                    "ts_code": "000001.SZ",
                    "name": "平安银行",
                    "trade_date": "20260630",
                    "current_close": 10.5,
                    "today_rank": 1,
                    "previous_rank": 2,
                    "rank_change": -1,
                    "best_rank": 1,
                    "total_score": 78.0,
                    "total_score_change": 3.5,
                    "selected_count_5d": 2,
                    "selected_count_10d": 4,
                    "consecutive_selected_days": 2,
                    "watch_status": "active_watch",
                    "watch_status_label": "正常观察",
                    "elder_score": 62.0,
                    "action_hint": "趋势尚可，等待回调",
                    "elder_reason": "节奏复核",
                    "weekly_trend": "改善",
                    "daily_pullback": "接近均线",
                    "force_signal": "转强",
                    "elder_ray_signal": "多头改善",
                }
            ]
        ),
    )
    store.upsert_dataframe(
        "external_position_snapshots",
        pd.DataFrame(
            [
                {
                    "id": "external-000001-20260630",
                    "platform": "模拟",
                    "account_name": "默认",
                    "snapshot_date": "20260630",
                    "ts_code": "000001.SZ",
                    "name": "平安银行",
                    "quantity": 100,
                    "cost_price": 10.0,
                    "current_price": 10.5,
                    "market_value": 1050.0,
                    "pnl": 50.0,
                    "pnl_pct": 0.05,
                    "risk_status": "normal",
                    "risk_status_cn": "正常",
                }
            ]
        ),
    )
    return store


def _settings(store: DuckDBStore) -> SimpleNamespace:
    return SimpleNamespace(
        duckdb_path=store.db_path,
        data_provider="akshare",
        real_universe_preset="full",
        akshare_sample_symbols="",
    )


def _write_lookback_status(
    tmp_path: Path,
    *,
    as_of_trade_date: str = "20260630",
    candidate_sample_count: int = 30,
    valid_sample_count: int = 24,
    total_score_group_summary: str = "综合分分组摘要",
    elder_review_summary: str = "埃尔德复核摘要",
    entry_zone_summary: str = "买入区间摘要",
    watchlist_summary: str = "观察池状态摘要",
    key_findings: str = "主要发现",
) -> Path:
    status_path = tmp_path / "lookback_status.json"
    status_path.write_text(
        json.dumps(
            {
                "status": "success",
                "as_of_trade_date": as_of_trade_date,
                "start_date": "20260601",
                "end_date": as_of_trade_date,
                "horizons": [1, 3, 5, 10, 20],
                "candidate_sample_count": candidate_sample_count,
                "valid_sample_count": valid_sample_count,
                "insufficient_forward_data_count": 6,
                "total_score_group_summary": total_score_group_summary,
                "elder_review_summary": elder_review_summary,
                "entry_zone_summary": entry_zone_summary,
                "watchlist_summary": watchlist_summary,
                "key_findings": key_findings,
                "data_quality_summary": "数据质量提示",
                "generated_report_path": "/tmp/lookback.xlsx",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return status_path


def _table_counts(store: DuckDBStore) -> dict[str, int]:
    tables = [
        "strategy_result",
        "factor_scores",
        "entry_zone_snapshots",
        "watchlist_daily_snapshots",
        "external_position_snapshots",
    ]
    return {table: len(store.read_table(table)) for table in tables}
